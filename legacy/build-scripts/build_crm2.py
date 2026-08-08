from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

FONT_NAME = "Arial"
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
NORMAL = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, size=10, bold=True)
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF")
GREEN_FONT = Font(name=FONT_NAME, size=10, color="008000")
GREY_FONT = Font(name=FONT_NAME, size=10, color="999999")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE5CD")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
GREY_FILL = PatternFill("solid", fgColor="EFEFEF")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
BLACK_FILL = PatternFill("solid", fgColor="D9D9D9")
DIVIDER_FILL = PatternFill("solid", fgColor="C9DAF8")

wb = Workbook()

def style_header_row(ws, row, ncols, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ============================================================
# SHARED CONFIG
# ============================================================
REPS = [
    {"name": "Tuna", "code": "TU", "channel": "Instagram", "default_source": "Instagram",
     "kpi": {"outreach": 100, "followups": 40, "conversations": 20, "kyc_started": 5, "kyc_complete": 3,
             "ftd": 1, "vip_transfers": 3, "checkins": 10}},
    {"name": "Chella", "code": "CH", "channel": "Discord / Community", "default_source": "Discord",
     "kpi": {"outreach": 50, "followups": 40, "conversations": 15, "kyc_started": 5, "kyc_complete": 3,
             "ftd": 2, "vip_transfers": 1, "checkins": 10}},
    {"name": "Moneyheist", "code": "MH", "channel": "Discord / Community", "default_source": "Discord",
     "kpi": {"outreach": 50, "followups": 40, "conversations": 15, "kyc_started": 5, "kyc_complete": 3,
             "ftd": 2, "vip_transfers": 1, "checkins": 10}},
    {"name": "Seb", "code": "SB", "channel": "SlotEssentials / Social", "default_source": "SlotEssentials",
     "kpi": {"outreach": 30, "followups": 40, "conversations": 15, "kyc_started": 5, "kyc_complete": 3,
             "ftd": 2, "vip_transfers": 1, "checkins": 10}},
    {"name": "Jordan", "code": "JD", "channel": "Discord + SlotEssentials", "default_source": "SlotEssentials",
     "kpi": {"outreach": 30, "followups": 40, "conversations": 15, "kyc_started": 5, "kyc_complete": 3,
             "ftd": 2, "vip_transfers": 1, "checkins": 10}},
    {"name": "Seanok", "code": "SK", "channel": "SlotEssentials", "default_source": "SlotEssentials",
     "kpi": {"outreach": 30, "followups": 40, "conversations": 15, "kyc_started": 5, "kyc_complete": 3,
             "ftd": 2, "vip_transfers": 1, "checkins": 10}},
]
REP_NAMES = [r["name"] for r in REPS]

SOURCE_CHANNELS = ["Instagram", "Discord", "Twitter", "Telegram", "SlotEssentials", "Other"]
STATUS_STAGES = ["New Lead", "Initial Contact", "Interested", "KYC Started", "KYC Complete",
                  "Deposit Pending", "First Deposit", "Active Week 1", "Active Week 2", "Active Week 3",
                  "Active Week 4", "Transfer to VIP", "Reactivation Queue", "Dead Lead"]
STATUS_NEXTACTION = {
    "New Lead": "Day 0: Initial Contact - Answer Questions, Help Register, Begin KYC",
    "Initial Contact": "Day 1: Check Account, Check KYC, Help Deposit",
    "Interested": "Check Account, Check KYC, Help Deposit",
    "KYC Started": "Follow up: Confirm KYC completion",
    "KYC Complete": "Help Deposit / Confirm Deposit Pending",
    "Deposit Pending": "Confirm Deposit, Resolve Issues",
    "First Deposit": "Day 3: Confirm Playing, Resolve Issues",
    "Active Week 1": "Day 7: Check Activity, Encourage Continued Play",
    "Active Week 2": "Day 14: Continue Relationship",
    "Active Week 3": "Continue Relationship, Monitor Activity",
    "Active Week 4": "Day 30 Decision: Transfer to VIP if active, else Reactivation Queue",
    "Transfer to VIP": "Handed to VIP team - no acquisition action",
    "Reactivation Queue": "Reactivation outreach - win them back",
    "Dead Lead": "No action - dead lead",
}
STATUS_OFFSET = {
    "New Lead": 1, "Initial Contact": 1, "Interested": 1, "KYC Started": 1, "KYC Complete": 1,
    "Deposit Pending": 1, "First Deposit": 3, "Active Week 1": 7, "Active Week 2": 7,
    "Active Week 3": 14, "Active Week 4": 30, "Transfer to VIP": "CLOSED",
    "Reactivation Queue": 3, "Dead Lead": "CLOSED",
}
KYC_STATUSES = ["Not Started", "Started", "Complete", "Failed"]
DEPOSIT_STATUSES = ["No", "Pending", "Yes"]

BOOK_ROWS = 300       # capacity per rep book
DAILY_TASK_ROWS = 150 # capacity per rep daily task queue
QUEUE_ROWS = 300      # company-wide queue capacity

print("config ok - reps:", REP_NAMES)

# ============================================================
# SHEET: Read Me
# ============================================================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = "Daily Gamba — Acquisition CRM (Pilot: Tuna)"
ws["B2"].font = TITLE_FONT
ws["B3"] = "This is a one-rep pilot so we can lock the mechanics before rolling out to the full team."
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color="666666")

rows = [
    ("", ""),
    ("What's in this pilot", "header"),
    ("Tuna - Book", "— His full player database. Every player he owns, one row each. Status drives everything else automatically."),
    ("Tuna - Daily Task", "— Auto-generated: exactly who he needs to contact today, ranked, pulled live from his Book. He works this top to bottom."),
    ("Tuna - Stats", "— His personal scoreboard: total list vs. active list, VIP transfers, wager."),
    ("Team & KPI Targets", "— His daily KPI targets (editable)."),
    ("", ""),
    ("How the automation works", "header"),
    ("Status (dropdown)", "— He sets this per player: New Lead -> Initial Contact -> Interested -> KYC Started -> KYC Complete -> Deposit Pending -> First Deposit -> Active Week 1-4 -> Transfer to VIP. Or Reactivation Queue / Dead Lead if they go cold."),
    ("Health", "— Auto-calculated. Green = on schedule, Yellow = follow-up due, Red = overdue 4+ days, Black = dead lead."),
    ("Next Action", "— Auto-calculated from Status. Tells him exactly what to do, no guessing."),
    ("Next Follow-Up Date", "— Auto-calculated from Status + Last Contact Date, using the Day 0/1/3/7/14/30 cadence."),
    ("Daily Task tab", "— Pulls in anyone whose Next Follow-Up Date is today or earlier. That's his whole to-do list."),
    ("", ""),
    ("Two things formulas can't do (need one small script)", "header"),
    ("Date Assigned", "— Ideally freezes automatically the moment a new player is added. In this pilot it's a manual date entry; I've got an Apps Script ready that automates it once we lock the structure."),
    ("Last Contact Date", "— Ideally auto-updates when he checks a player off the Daily Task list. Same story - manual for now, automatable with the same script."),
    ("", ""),
    ("Next step", "header"),
    ("Once this is confirmed", "— I'll replicate this exact structure for Chella, Moneyheist, Seb, Jordan, and Seanok, then build the Master Player DB, company-wide Daily Task Queue, Activity Log, and Executive Dashboard on top of all six."),
]
r = 5
for label, val in rows:
    if val == "header":
        ws.cell(row=r, column=2, value=label).font = SUBTITLE_FONT
        r += 1
        continue
    c1 = ws.cell(row=r, column=2, value=label)
    c1.font = BOLD
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws.cell(row=r, column=3, value=val)
    c2.font = NORMAL
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

autosize(ws, {"A": 3, "B": 26, "C": 95})
for col in ["D", "E", "F", "G"]:
    ws.column_dimensions[col].width = 12

print("readme ok")

# ============================================================
# HIDDEN SHEET: Lists
# ============================================================
lst = wb.create_sheet("Lists")
lst.sheet_state = "hidden"

def write_list(col, header, items):
    lst.cell(row=1, column=col, value=header)
    for i, v in enumerate(items, start=2):
        lst.cell(row=i, column=col, value=v)

write_list(1, "Reps", REP_NAMES)
write_list(2, "SourceChannels", SOURCE_CHANNELS)
write_list(4, "KYCStatuses", KYC_STATUSES)
write_list(5, "DepositStatuses", DEPOSIT_STATUSES)

# Status mapping table: col C=Status, D=NextAction, E=OffsetDays (or CLOSED)
lst.cell(row=1, column=3, value="Status")
lst.cell(row=1, column=4, value="NextAction")
lst.cell(row=1, column=5, value="OffsetDays")
for i, s in enumerate(STATUS_STAGES, start=2):
    lst.cell(row=i, column=3, value=s)
    lst.cell(row=i, column=4, value=STATUS_NEXTACTION[s])
    lst.cell(row=i, column=5, value=STATUS_OFFSET[s])

def col_range(col_idx, n_items):
    letter = get_column_letter(col_idx)
    return f"Lists!${letter}$2:${letter}${1+n_items}"

REPS_RANGE = col_range(1, len(REP_NAMES))
SOURCE_RANGE = col_range(2, len(SOURCE_CHANNELS))
STATUS_RANGE = col_range(3, len(STATUS_STAGES))
KYC_RANGE = col_range(4, len(KYC_STATUSES))
DEPOSIT_RANGE = col_range(5, len(DEPOSIT_STATUSES))
STATUS_TABLE = f"Lists!$C$2:$E${1+len(STATUS_STAGES)}"

print("lists ok")

# ============================================================
# SHEET: Team & KPI Targets
# ============================================================
ws = wb.create_sheet("Team & KPI Targets")
ws.sheet_view.showGridLines = False
ws["A1"] = "Team Roster & Daily KPI Targets"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Blue cells are editable inputs. Pilot includes Tuna only; other reps added once this structure is confirmed."
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

headers = ["Name", "Role / Channel", "Daily New\nOutreach", "Daily\nFollow-Ups", "Daily Meaningful\nConversations",
           "Daily KYC\nStarted", "Daily KYC\nCompleted", "Daily First\nDeposits", "Daily VIP\nTransfers",
           "Daily Active\nCheck-ins", "Daily CRM\nUpdates"]
HEADER_ROW = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header_row(ws, HEADER_ROW, len(headers), height=32)

team_rows = [
    ("Andis", "Manager - Sales Force Oversight (not on daily grind quota)", None, None, None, None, None, None, None, None, None),
    ("Tuna", "Instagram", 100, 40, 20, 5, 3, 1, 3, 10, "100%"),
]

r = HEADER_ROW + 1
TEAM_FIRST_ROW = r
for row in team_rows:
    for i, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) if i > 2 else Alignment(horizontal="left", vertical="center", wrap_text=True)
        if i >= 3:
            if val is None:
                cell.value = "-"
                cell.font = GREY_FONT
            else:
                cell.font = INPUT_FONT
        elif i == 1:
            cell.font = BOLD
        else:
            cell.font = NORMAL
    ws.row_dimensions[r].height = 30
    r += 1

LAST_TEAM_ROW = r - 1
autosize(ws, {"A": 13, "B": 42, "C": 11, "D": 11, "E": 13, "F": 10, "G": 10, "H": 10, "I": 10, "J": 10, "K": 10})
ws.freeze_panes = "A5"

print("team ok", TEAM_FIRST_ROW, LAST_TEAM_ROW)

# ============================================================
# FUNCTION: build a rep's Book tab
# ============================================================
BOOK_META = {}  # rep name -> dict with row/col info for reuse by Daily Task / Stats / Master

def build_book_sheet(rep):
    name = rep["name"]
    code = rep["code"]
    default_source = rep["default_source"]
    ws = wb.create_sheet(f"{name} - Book")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Book — Player Database"
    ws["A1"].font = TITLE_FONT

    headers = ["Player ID", "Player Name / Handle", "Source", "Roobet Username", "Status", "Health",
               "Priority\n(1=Urgent)", "Date\nAssigned", "Last Contact\nDate", "Next Follow-Up\nDue",
               "Next Action", "KYC Status", "Deposit Status", "Weighted\nWager", "VIP Ready", "Notes",
               "DaysSinceContact", "DueFlag", "DueRank"]
    HEADER_ROW = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=36)

    FIRST_ROW = HEADER_ROW + 1
    LAST_ROW = FIRST_ROW + BOOK_ROWS - 1

    for idx in range(BOOK_ROWS):
        r = FIRST_ROW + idx
        ws.cell(row=r, column=1, value=f'=IF($B{r}="","","{code}-"&TEXT(ROW()-{HEADER_ROW},"0000"))')
        ws.cell(row=r, column=3, value=default_source)

        # Health (F)
        f_health = (
            f'=IF($E{r}="","",'
            f'IF($E{r}="Dead Lead","⚫ Black",'
            f'IF($E{r}="Transfer to VIP","\U0001f7e2 Green",'
            f'IF($E{r}="Reactivation Queue","\U0001f534 Red",'
            f'IF($I{r}="","\U0001f7e1 Yellow",'
            f'IF(ISNUMBER($J{r}),'
            f'IF(TODAY()-$J{r}>3,"\U0001f534 Red",IF(TODAY()-$J{r}>=0,"\U0001f7e1 Yellow","\U0001f7e2 Green")),'
            f'"\U0001f7e2 Green"))))))'
        )
        ws.cell(row=r, column=6, value=f_health)

        # Priority (G)
        f_prio = (
            f'=IF($F{r}="","",IF($F{r}="\U0001f534 Red",1,IF($F{r}="\U0001f7e1 Yellow",2,'
            f'IF($F{r}="\U0001f7e2 Green",3,4))))'
        )
        ws.cell(row=r, column=7, value=f_prio)

        # Next Follow-Up (J)
        f_next = (
            f'=IF($E{r}="","",IF($I{r}="","",'
            f'IF(VLOOKUP($E{r},{STATUS_TABLE},3,FALSE)="CLOSED","N/A - Closed",'
            f'$I{r}+VLOOKUP($E{r},{STATUS_TABLE},3,FALSE))))'
        )
        ws.cell(row=r, column=10, value=f_next)

        # Next Action (K)
        f_action = f'=IF($E{r}="","",IFERROR(VLOOKUP($E{r},{STATUS_TABLE},2,FALSE),""))'
        ws.cell(row=r, column=11, value=f_action)

        # VIP Ready (O)
        f_vip = f'=IF($E{r}="","",IF(AND($E{r}="Active Week 4",$F{r}="\U0001f7e2 Green"),"Yes","No"))'
        ws.cell(row=r, column=15, value=f_vip)

        # Days Since Contact (Q, hidden helper, col 17)
        ws.cell(row=r, column=17, value=f'=IF($I{r}="","",TODAY()-$I{r})')
        # DueFlag (R, col 18)
        ws.cell(row=r, column=18, value=f'=IF(AND($B{r}<>"",ISNUMBER($J{r}),$J{r}<=TODAY()),1,0)')
        # DueRank (S, col 19)
        ws.cell(row=r, column=19, value=f'=IF($R{r}=1,SUM($R${FIRST_ROW}:$R{r}),"")')

        for c in range(1, 17):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        for c in (8, 9, 10):
            ws.cell(row=r, column=c).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=8).font = INPUT_FONT
        ws.cell(row=r, column=9).font = INPUT_FONT
        ws.cell(row=r, column=14).number_format = "$#,##0;($#,##0);-"

    dv_source = DataValidation(type="list", formula1=SOURCE_RANGE, allow_blank=True)
    dv_status = DataValidation(type="list", formula1=STATUS_RANGE, allow_blank=True)
    dv_kyc = DataValidation(type="list", formula1=KYC_RANGE, allow_blank=True)
    dv_dep = DataValidation(type="list", formula1=DEPOSIT_RANGE, allow_blank=True)
    for dv in (dv_source, dv_status, dv_kyc, dv_dep):
        ws.add_data_validation(dv)
    dv_source.add(f"C{FIRST_ROW}:C{LAST_ROW}")
    dv_status.add(f"E{FIRST_ROW}:E{LAST_ROW}")
    dv_kyc.add(f"L{FIRST_ROW}:L{LAST_ROW}")
    dv_dep.add(f"M{FIRST_ROW}:M{LAST_ROW}")

    rngh = f"F{FIRST_ROW}:F{LAST_ROW}"
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="\U0001f534 Red"'], fill=RED_FILL))
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="\U0001f7e1 Yellow"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="\U0001f7e2 Green"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="⚫ Black"'], fill=BLACK_FILL))

    for col in ("Q", "R", "S"):
        ws.column_dimensions[col].hidden = True

    autosize(ws, {"A": 10, "B": 20, "C": 13, "D": 16, "E": 16, "F": 11, "G": 9, "H": 11, "I": 12,
                  "J": 13, "K": 34, "L": 12, "M": 12, "N": 11, "O": 9, "P": 26})
    ws.freeze_panes = f"C{FIRST_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:S{LAST_ROW}"

    BOOK_META[name] = {"sheet": f"{name} - Book", "header_row": HEADER_ROW,
                        "first_row": FIRST_ROW, "last_row": LAST_ROW}
    print(f"book ok: {name}", FIRST_ROW, LAST_ROW)

for rep in REPS:
    if rep["name"] == "Tuna":
        build_book_sheet(rep)

# ============================================================
# FUNCTION: build a rep's Daily Task tab
# ============================================================
def build_daily_task_sheet(rep):
    name = rep["name"]
    meta = BOOK_META[name]
    BK = f"'{meta['sheet']}'"
    b_first, b_last = meta["first_row"], meta["last_row"]

    ws = wb.create_sheet(f"{name} - Daily Task")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Daily Task Queue"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Auto-pulled from his Book. Anyone overdue or due today, ranked. Work top to bottom - no thinking, just execute."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    headers = ["Done", "Match Row", "Player ID", "Player Name", "Due Date", "Owner", "Source",
               "Roobet Username", "Status", "Health", "Required Action", "Days Since\nLast Contact", "Notes"]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=32)

    FIRST_ROW = HEADER_ROW + 1
    LAST_ROW = FIRST_ROW + DAILY_TASK_ROWS - 1

    for idx in range(DAILY_TASK_ROWS):
        r = FIRST_ROW + idx
        priority = idx + 1
        ws.cell(row=r, column=1, value=False)
        ws.cell(row=r, column=2, value=f'=IFERROR(MATCH({priority},{BK}!$S${b_first}:$S${b_last},0)+{b_first}-1,"")')
        mref = f"$B{r}"
        ws.cell(row=r, column=3, value=f'=IF({mref}="","",INDEX({BK}!$A:$A,{mref}))')
        ws.cell(row=r, column=4, value=f'=IF({mref}="","",INDEX({BK}!$B:$B,{mref}))')
        ws.cell(row=r, column=5, value=f'=IF({mref}="","",INDEX({BK}!$J:$J,{mref}))')
        ws.cell(row=r, column=6, value=f'=IF({mref}="","","{name}")')
        ws.cell(row=r, column=7, value=f'=IF({mref}="","",INDEX({BK}!$C:$C,{mref}))')
        ws.cell(row=r, column=8, value=f'=IF({mref}="","",INDEX({BK}!$D:$D,{mref}))')
        ws.cell(row=r, column=9, value=f'=IF({mref}="","",INDEX({BK}!$E:$E,{mref}))')
        ws.cell(row=r, column=10, value=f'=IF({mref}="","",INDEX({BK}!$F:$F,{mref}))')
        ws.cell(row=r, column=11, value=f'=IF({mref}="","",INDEX({BK}!$K:$K,{mref}))')
        ws.cell(row=r, column=12, value=f'=IF({mref}="","",INDEX({BK}!$Q:$Q,{mref}))')
        ws.cell(row=r, column=13, value=f'=IF({mref}="","",INDEX({BK}!$P:$P,{mref}))')
        for c in range(1, 14):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"

    dv_done = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_done)
    dv_done.add(f"A{FIRST_ROW}:A{LAST_ROW}")

    rng = f"J{FIRST_ROW}:J{LAST_ROW}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$J{FIRST_ROW}="\U0001f534 Red"'], fill=RED_FILL))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$J{FIRST_ROW}="\U0001f7e1 Yellow"'], fill=YELLOW_FILL))

    ws.column_dimensions["B"].hidden = True
    autosize(ws, {"A": 8, "C": 10, "D": 20, "E": 12, "F": 10, "G": 12, "H": 16, "I": 16,
                  "J": 10, "K": 34, "L": 12, "M": 26})
    ws.freeze_panes = f"C{FIRST_ROW}"

    print(f"daily task ok: {name}", FIRST_ROW, LAST_ROW)

for rep in REPS:
    if rep["name"] == "Tuna":
        build_daily_task_sheet(rep)

# ============================================================
# FUNCTION: build a rep's Stats tab
# ============================================================
ACTIVE_STATUSES = ["Active Week 1", "Active Week 2", "Active Week 3", "Active Week 4"]

def build_stats_sheet(rep):
    name = rep["name"]
    meta = BOOK_META[name]
    BK = f"'{meta['sheet']}'"
    b_first, b_last = meta["first_row"], meta["last_row"]
    status_rng = f"{BK}!$E${b_first}:$E${b_last}"
    id_rng = f"{BK}!$B${b_first}:$B${b_last}"
    wager_rng = f"{BK}!$N${b_first}:$N${b_last}"
    dep_rng = f"{BK}!$M${b_first}:$M${b_last}"

    ws = wb.create_sheet(f"{name} - Stats")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Statistics"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Total List (all players ever assigned)"
    ws["A3"].font = SUBTITLE_FONT
    tl_headers = ["Number of Contacts", "Number of VIP Transfers", "Number of Active Players", "Total Wagered (current)"]
    for i, h in enumerate(tl_headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(tl_headers), height=28)
    ws.cell(row=5, column=1, value=f'=COUNTIF({id_rng},"<>")')
    ws.cell(row=5, column=2, value=f'=COUNTIF({status_rng},"Transfer to VIP")')
    active_terms = ",".join([f'"{s}"' for s in ACTIVE_STATUSES])
    ws.cell(row=5, column=3, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{active_terms}}}))')
    ws.cell(row=5, column=4, value=f'=SUM({wager_rng})')
    ws.cell(row=5, column=4).number_format = "$#,##0;($#,##0);-"
    for c in range(1, 5):
        ws.cell(row=5, column=c).border = BORDER
        ws.cell(row=5, column=c).font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
        ws.cell(row=5, column=c).alignment = Alignment(horizontal="center")

    ws["A7"] = "Active List (currently active players only)"
    ws["A7"].font = SUBTITLE_FONT
    al_headers = ["Number of Contacts", "Number of VIP Transfers", "Number of Active Players", "Number of FTDs"]
    for i, h in enumerate(al_headers, start=1):
        ws.cell(row=8, column=i, value=h)
    style_header_row(ws, 8, len(al_headers), height=28)
    ws.cell(row=9, column=1, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{active_terms}}}))')
    ws.cell(row=9, column=2, value=f'=COUNTIF({status_rng},"Transfer to VIP")')
    ws.cell(row=9, column=3, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{active_terms}}}))')
    fdterms = ",".join([f'"{s}"' for s in (["First Deposit"] + ACTIVE_STATUSES + ["Transfer to VIP"])])
    ws.cell(row=9, column=4, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{fdterms}}}))')
    for c in range(1, 5):
        ws.cell(row=9, column=c).border = BORDER
        ws.cell(row=9, column=c).font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
        ws.cell(row=9, column=c).alignment = Alignment(horizontal="center")

    ws["A11"] = "Note: wager figures are a running total from the Weighted Wager column in his Book. A dedicated monthly/YTD wager log can be added in Phase 2 once wager data is fed in regularly."
    ws["A11"].font = Font(name=FONT_NAME, size=8, italic=True, color="666666")
    ws["A11"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[11].height = 28

    autosize(ws, {"A": 24, "B": 24, "C": 24, "D": 22})
    print(f"stats ok: {name}")

for rep in REPS:
    if rep["name"] == "Tuna":
        build_stats_sheet(rep)

wb.save("Tuna_Pilot_CRM.xlsx")
print("SAVED OK")
