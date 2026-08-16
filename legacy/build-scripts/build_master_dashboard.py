from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
NORMAL = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, size=10, bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
BLACK_FILL = PatternFill("solid", fgColor="D9D9D9")

wb = Workbook()

def style_header_row(ws, row, ncols, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ============================================================
# CONFIG - add a rep's real Google Sheet ID here once their file exists.
# Only reps listed here get pulled into the Master file.
# ============================================================
# Daily targets are NOT stored here - they're pulled live from the roster's own
# Team & KPI Targets tab (see ROSTER import below), so there's exactly one place to
# edit a target (any rep's file) instead of keeping numbers in sync in two places.
REPS = [
    {"name": "Yuri", "sheet_id": "1_MRriw8lIFqUf6393GbCnTyzLwpAeLRmkIVhT6BQUZ8"},
    {"name": "Tuna", "sheet_id": "1E2DMtbb0qnQMj6PLoA3xPq59m3Gt0vnJl7UopBaMqhc"},
    {"name": "Plat", "sheet_id": "1Fk54o-UXm0eMEzi-5e1_yqmrFmyLaoRfMjLThRQRKoY"},
    {"name": "Chella", "sheet_id": "1RQVsxe270nT9Dw9GCvmcwlzuVL6D3QFNCmm1KpZOGp4"},
    {"name": "Moneyheist", "sheet_id": "1zjxFDteeU3YWlO9w9B92o-2ol3_yJiNnttovezeMnOg"},
    {"name": "Seb", "sheet_id": "1ZWfAr3sHXFGpNgieFKycyZq0GxTNMJTcAG668Pk15rg"},
    {"name": "Pricey", "sheet_id": "1ZU7EXQAHws9qptiNUTqPi6fKE1ALTdipfU-gSaNNUAk"},
    {"name": "Seanok", "sheet_id": "1e0krwNw2H4hns9MSex-7cGvwH7epyhkKIxvVfT9Svu4"},
    {"name": "Gwen", "sheet_id": "1PusLuNTxs7n2kCQJRCD1K0XnSwrpBZDaelz51bcEho0"},
    {"name": "Miko", "sheet_id": "1GElB1aCPsITtPtn-ebNIPW9fxj5czH-H8Tzvo0umY4Y"},
    {"name": "Concept", "sheet_id": "16XZ2V3L5SSrcuhjDDyxKYH5XX8HStG4sUWngBnRmDLo"},
    # Daily now has a real (low) quota - see her Team & KPI Targets tab.
    {"name": "Daily", "sheet_id": "1_5BmTcpK3SWziQWPleIrTcnfciVnNJS9prFkIbbHZzw", "has_kpi": True},
    {"name": "Prime", "sheet_id": "1aWOmLGt-ajmq4lGyc2FFA-JsJqrIBogv0Cj0NKE2gKQ", "has_kpi": True},
]

BOOK_ROWS = 200         # must match each rep file's Book capacity
ACTLOG_ROWS = 5000      # generous headroom for Activity Log pull
FTD_LIST_ROWS = 200     # must match each rep file's FTD List capacity (same as BOOK_ROWS)

# ============================================================
# SHEET: Read Me
# ============================================================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = "Daily Gamba — Master Dashboard"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Company-wide view, pulled live from each rep's own sheet. This file is read-only - edit a player on their rep's own Book tab, not here."
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color="666666")

rm_rows = [
    ("", ""),
    ("What's in this file", "header"),
    ("Master Player DB", "— Every player, every rep, combined into one view."),
    ("VIP Pipeline", "— Every player in the VIP fast-track or handed to the in-house VIP team, company-wide, with overdue check-ins flagged in red."),
    ("Overdue Follow-Ups", "— Every lead, any rep, that's sat untouched 24+ hours past its Next Follow-Up Due date - company-wide, all in red."),
    ("Executive Dashboard", "— Today / This Week / This Month numbers vs. daily targets, company-wide."),
    ("Team & KPI Targets", "— Every rep's daily targets in one table, pulled from their own files."),
    ("", ""),
    ("One-time setup", "header"),
    ("First time you open this", "— Google will ask you to \"Allow access\" for each rep's sheet it's pulling from (via a function called IMPORTRANGE). Click Allow once per rep and it's done permanently."),
    ("Adding a new rep", "— Once a new rep's file exists, tell Claude their Google Sheet URL and it gets added here automatically - no other changes needed."),
]
r = 5
for label, val in rm_rows:
    if val == "header":
        ws.cell(row=r, column=2, value=label).font = SUBTITLE_FONT
        r += 1
        continue
    c1 = ws.cell(row=r, column=2, value=label)
    c1.font = BOLD
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws.cell(row=r, column=3, value=val)
    c2.font = NORMAL
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1
autosize(ws, {"A": 3, "B": 24, "C": 90})
print("readme ok")

# ============================================================
# HIDDEN IMPORT TABS - one pair per rep, raw IMPORTRANGE pulls.
# Everything else in this workbook reads from these, never straight from
# the rep's file, so there's exactly one place doing the cross-file pull.
# ============================================================
IMPORT_META = {}
for rep in REPS:
    name = rep["name"]
    sid = rep["sheet_id"]
    has_kpi = rep.get("has_kpi", True)

    book_ws = wb.create_sheet(f"_Import {name} Book")
    book_ws["A1"] = f'=IMPORTRANGE("{sid}", "\'{name} - Book\'!A4:O{3 + BOOK_ROWS}")'

    log_ws = wb.create_sheet(f"_Import {name} ActivityLog")
    log_ws["A1"] = f'=IMPORTRANGE("{sid}", "\'{name} - Activity Log\'!A5:E{4 + ACTLOG_ROWS}")'

    # VIP Team column (AC on the rep's Book) isn't contiguous with A:O, so it gets its
    # own small import - same pattern as Book/ActivityLog above. Pulls Player Handle
    # (B) alongside it so this hidden tab is self-explanatory to read on its own -
    # col A = handle for reference, col B = the actual Yes/No/blank value.
    # Columns: A=Handle, B=Transferred to VIP Team value, C=VIP Transfer Date (W - the
    # in-house team hand-off date), D=VIP Transferred fast-track date (AA - the pre-
    # deposit funnel Status date). Used together on the VIP Pipeline as a single
    # "Transferred Date" sort key (whichever of the two is actually set).
    vipteam_ws = wb.create_sheet(f"_Import {name} VIPTeam")
    vipteam_ws["A1"] = (
        f'={{IMPORTRANGE("{sid}", "\'{name} - Book\'!B4:B{3 + BOOK_ROWS}"),'
        f'IMPORTRANGE("{sid}", "\'{name} - Book\'!AC4:AC{3 + BOOK_ROWS}"),'
        f'IMPORTRANGE("{sid}", "\'{name} - Book\'!W4:W{3 + BOOK_ROWS}"),'
        f'IMPORTRANGE("{sid}", "\'{name} - Book\'!AA4:AA{3 + BOOK_ROWS}")}}'
    )

    # FTD List pull - same shape as Book/ActivityLog, feeds the combined Master FTD List
    # tab below. Handle, Roobet Username, FTD Date, Weighted Wager, Total Wager (manual).
    ftd_list_ws = wb.create_sheet(f"_Import {name} FtdList")
    ftd_list_ws["A1"] = f'=IMPORTRANGE("{sid}", "\'{name} - FTD List\'!A5:E{4 + FTD_LIST_ROWS}")'

    kpi_sheet_name = None
    if has_kpi:
        # Each rep's Team & KPI Targets tab is personal to them - just their own single
        # row (row 5), not a shared roster view - so a plain per-rep import is correct here.
        kpi_ws = wb.create_sheet(f"_Import {name} KPI")
        kpi_ws["A1"] = f'=IMPORTRANGE("{sid}", "\'Team & KPI Targets\'!A5:G5")'
        kpi_sheet_name = f"'_Import {name} KPI'"

    IMPORT_META[name] = {
        "book_sheet": f"'_Import {name} Book'",
        "book_rows": BOOK_ROWS,
        "log_sheet": f"'_Import {name} ActivityLog'",
        "log_rows": ACTLOG_ROWS,
        "vipteam_sheet": f"'_Import {name} VIPTeam'",
        "ftd_list_sheet": f"'_Import {name} FtdList'",
        "ftd_list_rows": FTD_LIST_ROWS,
        "kpi_sheet": kpi_sheet_name,
        "has_kpi": has_kpi,
    }
print("import tabs ok:", list(IMPORT_META.keys()))

# ============================================================
# SHEET: Team & KPI Targets (company-wide, pulled from each rep file)
# ============================================================
ws = wb.create_sheet("Team & KPI Targets")
ws.sheet_view.showGridLines = False
ws["A1"] = "Team Roster & Daily KPI Targets — All Reps"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Pulled live from each rep's own Team & KPI Targets tab (personal to them). "
            "To change a target, edit it on that rep's own file - it updates here automatically.")
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

headers = ["Name", "Role", "Platforms", "Daily Outreach\n(static, not tracked)",
           "Active Leads\nTarget (must log)", "VIP Transfers\nTarget", "Daily FTD\nTarget"]
HEADER_ROW = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header_row(ws, HEADER_ROW, len(headers), height=32)

r = HEADER_ROW + 1
TEAM_FIRST_ROW = r
for rep in REPS:
    name = rep["name"]
    if not rep.get("has_kpi", True):
        continue
    KP = IMPORT_META[name]["kpi_sheet"]
    for c in range(1, 8):
        col_letter = get_column_letter(c)
        ws.cell(row=r, column=c, value=f'=IFERROR({KP}!{col_letter}1,"")')
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = NORMAL if c <= 3 else Font(name=FONT_NAME, size=10, color=NAVY)
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="left" if c <= 3 else "center", vertical="center")
    ws.row_dimensions[r].height = 26
    r += 1
TEAM_LAST_ROW = r - 1
autosize(ws, {"A": 13, "B": 26, "C": 24, "D": 12, "E": 12, "F": 12, "G": 10})
ws.freeze_panes = "A5"
print("team kpi ok:", TEAM_FIRST_ROW, TEAM_LAST_ROW)

# ============================================================
# SHEET: Master Player DB
# ============================================================
ws = wb.create_sheet("Master Player DB")
ws.sheet_view.showGridLines = False
ws["A1"] = "Master Player Database - All Reps Combined"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Live-pulled from every rep's Book tab, auto-compacted - blank capacity rows are "
            "stripped out so every rep's real leads sit back-to-back with no scrolling through "
            "empty space between reps. Read-only - to change a player's info, edit it on that "
            "rep's own Book tab, not here.")
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

headers = ["Rep", "Player ID", "Player Handle", "Source", "Roobet Username", "Status", "Health",
           "Date Assigned", "Last Contact Date", "Next Follow-Up", "Next Action", "KYC Status",
           "Deposit Status", "Weighted Wager", "VIP Ready", "Transferred to\nVIP Team"]
HEADER_ROW = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header_row(ws, HEADER_ROW, len(headers), height=32)

FIRST_ROW = HEADER_ROW + 1

# One combined array literal: each rep contributes a [RepLabel | Book A:F | Book H:O |
# VIPTeam Yes/No] block (BOOK_ROWS rows tall, columns pre-arranged in final output order
# with Priority (column G) simply skipped via the A:F / H:O split), stacked vertically
# across every rep. FILTER() then strips blank-handle rows using a parallel condition
# array built from the same per-rep Handle ranges - deliberately NOT QUERY, since QUERY's
# column-type auto-detection over array literals is unreliable when a column samples as
# blank (throws "NO_COLUMN" errors) - FILTER works purely positionally with no type
# inference, so it doesn't have that failure mode. This packs every rep's real leads
# together with zero blank rows in between, instead of the old approach of reserving a
# fixed BOOK_ROWS-row block per rep regardless of how many were filled (which is why
# Plat's one lead used to be buried 200 rows down under Tuna's empty capacity).
rep_blocks = []
cond_blocks = []
for rep in REPS:
    name = rep["name"]
    BK = IMPORT_META[name]["book_sheet"]
    VT = IMPORT_META[name]["vipteam_sheet"]
    rows = IMPORT_META[name]["book_rows"]
    rep_blocks.append(
        f'IF({BK}!B1:B{rows}="","","{name}"), {BK}!A1:F{rows}, {BK}!H1:O{rows}, '
        f'IFERROR({VT}!B1:B{rows},"")'
    )
    cond_blocks.append(f'{BK}!B1:B{rows}<>""')
combined = "; ".join(rep_blocks)
conditions = "; ".join(cond_blocks)

ws.cell(row=FIRST_ROW, column=1, value=(
    f'=IFERROR(FILTER({{{combined}}}, {{{conditions}}}), "")'
))

# Generous headroom for number formats + conditional formatting + autofilter, since the
# QUERY's actual spill size is dynamic and grows as reps log more leads (up to every rep
# hitting full BOOK_ROWS capacity at once).
MAX_ROWS = BOOK_ROWS * max(len(REPS), 1)
LAST_ROW = FIRST_ROW + MAX_ROWS - 1

for c in (8, 9, 10):  # Date Assigned, Last Contact Date, Next Follow-Up
    for rr in range(FIRST_ROW, LAST_ROW + 1):
        ws.cell(row=rr, column=c).number_format = "yyyy-mm-dd"
for rr in range(FIRST_ROW, LAST_ROW + 1):
    ws.cell(row=rr, column=14).number_format = "$#,##0;($#,##0);-"

rngh = f"G{FIRST_ROW}:G{LAST_ROW}"
ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$G{FIRST_ROW}="Red"'], fill=RED_FILL))
ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$G{FIRST_ROW}="Yellow"'], fill=YELLOW_FILL))
ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$G{FIRST_ROW}="Green"'], fill=GREEN_FILL))
ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$G{FIRST_ROW}="Black"'], fill=BLACK_FILL))

autosize(ws, {"A": 12, "B": 10, "C": 20, "D": 13, "E": 16, "F": 16, "G": 9, "H": 11,
              "I": 12, "J": 13, "K": 34, "L": 12, "M": 12, "N": 11, "O": 9, "P": 15})
ws.freeze_panes = f"C{FIRST_ROW}"
ws.auto_filter.ref = f"A{HEADER_ROW}:P{LAST_ROW}"
print("master db ok:", FIRST_ROW, LAST_ROW)

# ============================================================
# SHEET: Master FTD List - every First Time Depositor across every rep, one combined
# list. Same compaction pattern as Master Player DB above: FILTER strips blank-handle
# rows so everyone sits back-to-back with no empty capacity between reps. Weighted
# Wager only - no Total Wager column here. Total Wager is a per-player manual figure
# that only makes sense on the rep's own FTD List where they update it; the one number
# that matters company-wide is the combined Weighted Wager total, shown live below.
# ============================================================
ws = wb.create_sheet("Master FTD List")
ws.sheet_view.showGridLines = False
ws["A1"] = "Master FTD List - First Time Depositors, All Reps Combined"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Live-pulled from every rep's own FTD List tab, auto-compacted. Read-only - to "
            "add a player or update their Total Wager, do it on that rep's own FTD List tab, "
            "not here. A row highlighted red on a rep's own tab (stagnant wager) won't show "
            "that color here - check the rep's own FTD List to see who needs a check-in.")
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 28

ws.cell(row=3, column=1, value="Total Weighted Wager (all reps, current):")
ws.cell(row=3, column=1).font = BOLD
ws.cell(row=3, column=1).alignment = Alignment(horizontal="right")

ftd_headers = ["Rep", "Player Handle", "Roobet Username", "FTD Date", "Weighted Wager"]
FTD_HEADER_ROW = 4
for i, h in enumerate(ftd_headers, start=1):
    ws.cell(row=FTD_HEADER_ROW, column=i, value=h)
style_header_row(ws, FTD_HEADER_ROW, len(ftd_headers), height=32)

FTD_FIRST_ROW = FTD_HEADER_ROW + 1

ftd_rep_blocks = []
ftd_cond_blocks = []
for rep in REPS:
    name = rep["name"]
    FL = IMPORT_META[name]["ftd_list_sheet"]
    rows = IMPORT_META[name]["ftd_list_rows"]
    # Every rep's block and condition is individually wrapped in IFERROR - this is a
    # resilience fix. FILTER() fails its ENTIRE result if any single cell in any stacked
    # condition array errors out, so without this, one rep's broken source data (a row
    # deleted by hand on their FTD List tab, IMPORTRANGE not yet authorized, etc.) blanks
    # out every other rep's real entries too. With IFERROR here, a bad cell just drops
    # out of that one rep's block instead of nuking the combined list for everyone.
    ftd_rep_blocks.append(
        f'IFERROR(IF({FL}!A1:A{rows}="","","{name}"),""), IFERROR({FL}!A1:D{rows},"")'
    )
    ftd_cond_blocks.append(f'IFERROR({FL}!A1:A{rows}<>"",FALSE)')
ftd_combined = "; ".join(ftd_rep_blocks)
ftd_conditions = "; ".join(ftd_cond_blocks)

ws.cell(row=FTD_FIRST_ROW, column=1, value=(
    f'=IFERROR(FILTER({{{ftd_combined}}}, {{{ftd_conditions}}}), "")'
))

FTD_MAX_ROWS = FTD_LIST_ROWS * max(len(REPS), 1)
FTD_LAST_ROW = FTD_FIRST_ROW + FTD_MAX_ROWS - 1

ws.cell(row=3, column=2, value=f'=SUM(E{FTD_FIRST_ROW}:E{FTD_LAST_ROW})')
ws.cell(row=3, column=2).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
ws.cell(row=3, column=2).number_format = "$#,##0;($#,##0);-"

for rr in range(FTD_FIRST_ROW, FTD_LAST_ROW + 1):
    ws.cell(row=rr, column=4).number_format = "yyyy-mm-dd"
    ws.cell(row=rr, column=5).number_format = "$#,##0;($#,##0);-"

autosize(ws, {"A": 12, "B": 20, "C": 20, "D": 13, "E": 15})
ws.freeze_panes = f"B{FTD_FIRST_ROW}"
ws.auto_filter.ref = f"A{FTD_HEADER_ROW}:E{FTD_LAST_ROW}"
print("master ftd list ok:", FTD_FIRST_ROW, FTD_LAST_ROW)

# ============================================================
# SHEET: VIP Pipeline
# ============================================================
# Every player currently in EITHER VIP pipeline, company-wide: Status = "VIP Transferred"
# (referral-code signup, pre-deposit Day 1/2/3 fast-track) OR Transferred to VIP Team =
# "Yes" (handed to the in-house VIP team, Day 1/7/14 check-in cadence). Both cadences
# already feed into each rep's own Next Follow-Up Due / Next Action columns on their
# Book, so this view just reuses that existing math instead of re-deriving it - anyone
# whose Next Follow-Up Due is today or earlier is highlighted red as overdue.
ws = wb.create_sheet("VIP Pipeline")
ws.sheet_view.showGridLines = False
ws["A1"] = "VIP Pipeline - All Reps Combined"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Every player currently being handled by the in-house VIP team, or in the pre-deposit "
            "VIP fast-track, across every rep. Read-only. Rows in red are overdue for their next "
            "check-in - go to that rep's own Daily Task to act on it.")
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

vip_headers = ["Rep", "Player Handle", "Source", "Roobet Username", "Status", "Health",
               "Transferred to\nVIP Team", "Transferred\nDate", "Last Contact Date",
               "Next Follow-Up Due", "Next Action"]
VIP_HEADER_ROW = 4
for i, h in enumerate(vip_headers, start=1):
    ws.cell(row=VIP_HEADER_ROW, column=i, value=h)
style_header_row(ws, VIP_HEADER_ROW, len(vip_headers), height=32)

VIP_FIRST_ROW = VIP_HEADER_ROW + 1

vip_rep_blocks = []
vip_cond_blocks = []
for rep in REPS:
    name = rep["name"]
    BK = IMPORT_META[name]["book_sheet"]
    VT = IMPORT_META[name]["vipteam_sheet"]
    rows = IMPORT_META[name]["book_rows"]
    # Every block/condition is IFERROR-wrapped (same resilience fix as the Master FTD
    # List) so one rep's bad source data can't blank out everyone else's rows. The
    # condition also now requires a real Player Handle AND a real Roobet Username -
    # without this, a row whose handle got cleared but still has a stray Status/VIP Team
    # value left behind shows up as a blank "ghost" row, and a player marked VIP
    # Transferred / Transferred to VIP Team before ever getting a Roobet Username showed
    # up with no way to identify their account.
    # Transferred Date = whichever of VIP Transfer Date (actual team hand-off, VT col C)
    # or VIP Transferred fast-track date (pre-deposit funnel Status, VT col D) is set -
    # used purely as a "most recent activity" sort key so the list can be ordered newest
    # first instead of sitting in arbitrary Book row order.
    vip_rep_blocks.append(
        f'IFERROR(IF({BK}!B1:B{rows}="","","{name}"),""), IFERROR({BK}!B1:F{rows},""), '
        f'IFERROR(IFERROR({VT}!B1:B{rows},""),""), '
        f'IFERROR(IF(IFERROR({VT}!C1:C{rows},"")<>"",IFERROR({VT}!C1:C{rows},""),IFERROR({VT}!D1:D{rows},"")),""), '
        f'IFERROR({BK}!I1:K{rows},"")'
    )
    vip_cond_blocks.append(
        f'IFERROR((({BK}!E1:E{rows}="VIP Transferred")+(IFERROR({VT}!B1:B{rows},"")="Yes"))'
        f'*({BK}!B1:B{rows}<>"")*({BK}!D1:D{rows}<>""),FALSE)'
    )
vip_combined = "; ".join(vip_rep_blocks)
vip_conditions = "; ".join(vip_cond_blocks)

# Sorted newest-transferred-first (column 8 = Transferred Date, descending) instead of
# sitting in whatever order the underlying Books happen to list players.
ws.cell(row=VIP_FIRST_ROW, column=1, value=(
    f'=SORT(IFERROR(FILTER({{{vip_combined}}}, {{{vip_conditions}}}), ""), 8, FALSE)'
))

VIP_MAX_ROWS = BOOK_ROWS * max(len(REPS), 1)
VIP_LAST_ROW = VIP_FIRST_ROW + VIP_MAX_ROWS - 1

for rr in range(VIP_FIRST_ROW, VIP_LAST_ROW + 1):
    ws.cell(row=rr, column=8).number_format = "yyyy-mm-dd"   # Transferred Date
    ws.cell(row=rr, column=9).number_format = "yyyy-mm-dd"   # Last Contact Date
    ws.cell(row=rr, column=10).number_format = "yyyy-mm-dd"  # Next Follow-Up Due

overdue_rng = f"A{VIP_FIRST_ROW}:K{VIP_LAST_ROW}"
# Red only while they're still in the pre-deposit VIP Transferred stage and overdue for a
# check-in - once Status moves to First Deposit or Active, the urgency signal moves to
# the FTD List's stagnant-wager highlight instead, so this stops flagging automatically
# rather than staying red forever even after they've converted.
overdue_formula = f'AND($E{VIP_FIRST_ROW}="VIP Transferred",$J{VIP_FIRST_ROW}<>"",$J{VIP_FIRST_ROW}<=TODAY())'
ws.conditional_formatting.add(overdue_rng, FormulaRule(formula=[overdue_formula], fill=RED_FILL))

autosize(ws, {"A": 12, "B": 20, "C": 14, "D": 16, "E": 15, "F": 9, "G": 15,
              "H": 14, "I": 14, "J": 15, "K": 40})
ws.freeze_panes = f"B{VIP_FIRST_ROW}"
ws.auto_filter.ref = f"A{VIP_HEADER_ROW}:K{VIP_LAST_ROW}"
print("vip pipeline ok:", VIP_FIRST_ROW, VIP_LAST_ROW)

# ============================================================
# SHEET: Overdue Follow-Ups
# ============================================================
# Company-wide view of every lead, any rep, that's been sitting untouched 24+ hours -
# Next Follow-Up Due is strictly before today, meaning it was already due yesterday (or
# earlier) and nobody has logged a follow-up since. Same definition used for the red
# row-highlight on each rep's own Daily Task - this is just that same signal rolled up
# so it's visible without opening every rep's file individually.
ws = wb.create_sheet("Overdue Follow-Ups")
ws.sheet_view.showGridLines = False
ws["A1"] = "Overdue Follow-Ups - All Reps Combined"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Every lead, any rep, whose Next Follow-Up Due date has already passed and hasn't been "
            "acted on since. Read-only - go to that rep's own Daily Task to work it.")
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

overdue_headers = ["Rep", "Player Handle", "Status", "Health",
                    "Last Contact Date", "Next Follow-Up Due", "Next Action"]
OVERDUE_HEADER_ROW = 4
for i, h in enumerate(overdue_headers, start=1):
    ws.cell(row=OVERDUE_HEADER_ROW, column=i, value=h)
style_header_row(ws, OVERDUE_HEADER_ROW, len(overdue_headers), height=32)

OVERDUE_FIRST_ROW = OVERDUE_HEADER_ROW + 1

overdue_rep_blocks = []
overdue_cond_blocks = []
for rep in REPS:
    name = rep["name"]
    BK = IMPORT_META[name]["book_sheet"]
    rows = IMPORT_META[name]["book_rows"]
    overdue_rep_blocks.append(
        f'IF({BK}!B1:B{rows}="","","{name}"), {BK}!B1:B{rows}, {BK}!E1:F{rows}, {BK}!I1:K{rows}'
    )
    overdue_cond_blocks.append(
        f'(({BK}!J1:J{rows}<>"")*({BK}!J1:J{rows}<TODAY()))>0'
    )
overdue_combined = "; ".join(overdue_rep_blocks)
overdue_conditions = "; ".join(overdue_cond_blocks)

ws.cell(row=OVERDUE_FIRST_ROW, column=1, value=(
    f'=IFERROR(FILTER({{{overdue_combined}}}, {{{overdue_conditions}}}), "")'
))

OVERDUE_MAX_ROWS = BOOK_ROWS * max(len(REPS), 1)
OVERDUE_LAST_ROW = OVERDUE_FIRST_ROW + OVERDUE_MAX_ROWS - 1

for rr in range(OVERDUE_FIRST_ROW, OVERDUE_LAST_ROW + 1):
    ws.cell(row=rr, column=5).number_format = "yyyy-mm-dd"   # Last Contact Date
    ws.cell(row=rr, column=6).number_format = "yyyy-mm-dd"   # Next Follow-Up Due

overdue_all_rng = f"A{OVERDUE_FIRST_ROW}:G{OVERDUE_LAST_ROW}"
overdue_all_formula = f'$B{OVERDUE_FIRST_ROW}<>""'
ws.conditional_formatting.add(overdue_all_rng, FormulaRule(formula=[overdue_all_formula], fill=RED_FILL))

autosize(ws, {"A": 12, "B": 20, "C": 15, "D": 9, "E": 15, "F": 15, "G": 40})
ws.freeze_panes = f"B{OVERDUE_FIRST_ROW}"
ws.auto_filter.ref = f"A{OVERDUE_HEADER_ROW}:G{OVERDUE_LAST_ROW}"
print("overdue follow-ups ok:", OVERDUE_FIRST_ROW, OVERDUE_LAST_ROW)

# ============================================================
# SHEET: Executive Dashboard
# ============================================================
ws = wb.create_sheet("Executive Dashboard")
ws.sheet_view.showGridLines = False
ws["A1"] = "Executive Dashboard - Team Overview"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Live-pulled from each rep's Activity Log. Read-only - nothing to edit here."
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

# Pulls each rep's target from the Master's own Team & KPI Targets sheet (built above)
# by name - single-sourced, so it always matches whatever's on that rep's own file.
def target_lookup(rep_name, target_col_letter):
    return (
        f'IFERROR(INDEX(\'Team & KPI Targets\'!${target_col_letter}${TEAM_FIRST_ROW}:'
        f'${target_col_letter}${TEAM_LAST_ROW},MATCH("{rep_name}",'
        f'\'Team & KPI Targets\'!$A${TEAM_FIRST_ROW}:$A${TEAM_LAST_ROW},0)),"TBD")'
    )

# Row layout is precomputed up front (rather than incrementing as each block is written)
# so the Quick Overview block at the very top can reference the This Week/This Month
# table's rows before that table is actually written further down.
_n_reps = max(len(REPS), 1)
QO_TITLE_ROW = 4
QO_PERIODS = ["This Week", "This Month", "All Time"]
# QO_TITLE_ROW (block titles) + 3 data rows + 1 blank spacer + 1 "Today vs. Target"
# subtitle row = HEADER_ROW_1.
HEADER_ROW_1 = QO_TITLE_ROW + len(QO_PERIODS) + 3
FIRST_ROW_1 = HEADER_ROW_1 + 1
LAST_ROW_1 = FIRST_ROW_1 + _n_reps - 1
TOTAL_ROW_1 = LAST_ROW_1 + 1
HEADER_ROW_2 = TOTAL_ROW_1 + 3
FIRST_ROW_2 = HEADER_ROW_2 + 1
LAST_ROW_2 = FIRST_ROW_2 + _n_reps - 1

# ------------------------------------------------------------
# Quick Overview - two separate small blocks, VIP Transfers and FTD's side by side, each
# its own vertical This Week / This Month / All Time list (same style as a simple metrics
# card). Sits right at the top so it's the first thing seen on open. Week/Month sum
# straight from the This Week/This Month table below (single-sourced, always matches);
# All Time is its own COUNTIF sweep across every rep's full Activity Log history since no
# other block on this sheet tracks a running lifetime total.
# ------------------------------------------------------------
vip_alltime_terms = []
ftd_alltime_terms = []
for rep in REPS:
    name = rep["name"]
    LOG_qo = IMPORT_META[name]["log_sheet"]
    type_rng_qo = f"{LOG_qo}!$C$1:$C${IMPORT_META[name]['log_rows']}"
    vip_alltime_terms.append(f'COUNTIF({type_rng_qo},"VIP Transfer")+COUNTIF({type_rng_qo},"*-> VIP Transferred")')
    ftd_alltime_terms.append(f'COUNTIF({type_rng_qo},"*-> First Deposit")')

qo_blocks = [
    ("VIP Transfers", 1, [f"C{FIRST_ROW_2}:C{LAST_ROW_2}", f"F{FIRST_ROW_2}:F{LAST_ROW_2}", "+".join(vip_alltime_terms)]),
    ("FTD's", 4, [f"D{FIRST_ROW_2}:D{LAST_ROW_2}", f"G{FIRST_ROW_2}:G{LAST_ROW_2}", "+".join(ftd_alltime_terms)]),
]
for title, start_col, exprs in qo_blocks:
    label_col = start_col
    value_col = start_col + 1
    ws.cell(row=QO_TITLE_ROW, column=label_col, value=title)
    ws.cell(row=QO_TITLE_ROW, column=label_col).font = SUBTITLE_FONT
    for i, period in enumerate(QO_PERIODS):
        row = QO_TITLE_ROW + 1 + i
        expr = exprs[i]
        value_formula = f"=SUM({expr})" if i < 2 else f"={expr}"
        ws.cell(row=row, column=label_col, value=period)
        ws.cell(row=row, column=label_col).font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        ws.cell(row=row, column=label_col).fill = PatternFill("solid", fgColor="1F3864")
        ws.cell(row=row, column=label_col).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=row, column=value_col, value=value_formula)
        ws.cell(row=row, column=value_col).font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
        ws.cell(row=row, column=value_col).alignment = Alignment(horizontal="center")
        for c in (label_col, value_col):
            ws.cell(row=row, column=c).border = BORDER
        ws.row_dimensions[row].height = 22

ws.cell(row=HEADER_ROW_1 - 1, column=1, value="Today vs. Target")
ws.cell(row=HEADER_ROW_1 - 1, column=1).font = SUBTITLE_FONT
today_headers = ["Rep", "Active Leads Today", "Active Leads Target",
                  "VIP Transfers Today", "VIP Transfers Target", "FTD Today", "FTD Target"]
for i, h in enumerate(today_headers, start=1):
    ws.cell(row=HEADER_ROW_1, column=i, value=h)
style_header_row(ws, HEADER_ROW_1, len(today_headers), height=28)

r = HEADER_ROW_1 + 1
FIRST_ROW_1 = r
for rep in REPS:
    name = rep["name"]
    LOG = IMPORT_META[name]["log_sheet"]
    log_rows = IMPORT_META[name]["log_rows"]
    date_rng = f"{LOG}!$A$1:$A${log_rows}"
    type_rng = f"{LOG}!$C$1:$C${log_rows}"
    # VIP Transfers = handed to the in-house VIP team (logged as "VIP Transfer") OR a
    # Status change into "VIP Transferred" (logged generically as "Status Change: X -> VIP
    # Transferred") - matches the definition used on each rep's own Stats tab.
    vip_formula = (
        f'=COUNTIFS({date_rng},TODAY(),{type_rng},"VIP Transfer")'
        f'+COUNTIFS({date_rng},TODAY(),{type_rng},"*-> VIP Transferred")'
    )
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=1).font = BOLD
    # Active Leads counts by CURRENT status (excludes Dead Lead and Potential Lead), not
    # by the historical "Outreach" Activity Log event - matches the same definition used
    # on each rep's own Stats tab. Pulled from the Book import tab (columns E/H within its
    # A:O slice) since that's the only place that reflects a lead's LIVE status.
    BK_today = IMPORT_META[name]["book_sheet"]
    book_rows_today = IMPORT_META[name]["book_rows"]
    status_rng_today = f"{BK_today}!$E$1:$E${book_rows_today}"
    dateassigned_rng_today = f"{BK_today}!$H$1:$H${book_rows_today}"
    ws.cell(row=r, column=2, value=(
        f'=COUNTIFS({dateassigned_rng_today},TODAY(),'
        f'{status_rng_today},"<>Dead Lead",{status_rng_today},"<>Potential Lead")'
    ))
    ws.cell(row=r, column=3, value=f'={target_lookup(name, "E")}')
    ws.cell(row=r, column=4, value=vip_formula)
    ws.cell(row=r, column=5, value=f'={target_lookup(name, "F")}')
    ws.cell(row=r, column=6, value=f'=COUNTIFS({date_rng},TODAY(),{type_rng},"*-> First Deposit")')
    ws.cell(row=r, column=7, value=f'={target_lookup(name, "G")}')
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER
        if c > 1:
            ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=11, color=NAVY)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    r += 1
LAST_ROW_1 = r - 1

ws.cell(row=r, column=1, value="COMPANY TOTAL")
ws.cell(row=r, column=1).font = BOLD
for c in (2, 3, 4, 5, 6, 7):
    col_letter = get_column_letter(c)
    ws.cell(row=r, column=c, value=f'=SUM({col_letter}{FIRST_ROW_1}:{col_letter}{LAST_ROW_1})')
    ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
    ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=c).border = BORDER
TOTAL_ROW_1 = r

for actual_col, target_col in [(2, 3), (4, 5), (6, 7)]:
    al = get_column_letter(actual_col)
    tl = get_column_letter(target_col)
    rng = f"{al}{FIRST_ROW_1}:{al}{TOTAL_ROW_1}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${al}{FIRST_ROW_1}>=${tl}{FIRST_ROW_1}'], fill=GREEN_FILL))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${al}{FIRST_ROW_1}<${tl}{FIRST_ROW_1}'], fill=YELLOW_FILL))

r2 = TOTAL_ROW_1 + 2
ws.cell(row=r2, column=1, value="This Week / This Month")
ws.cell(row=r2, column=1).font = SUBTITLE_FONT
wm_headers = ["Rep", "Active Leads (Week)", "VIP Transfers (Week)", "FTD (Week)",
              "Active Leads (Month)", "VIP Transfers (Month)", "FTD (Month)"]
HEADER_ROW_2 = r2 + 1
for i, h in enumerate(wm_headers, start=1):
    ws.cell(row=HEADER_ROW_2, column=i, value=h)
style_header_row(ws, HEADER_ROW_2, len(wm_headers), height=28)

week_start = "(TODAY()-WEEKDAY(TODAY(),3))"
month_start = "DATE(YEAR(TODAY()),MONTH(TODAY()),1)"
r = HEADER_ROW_2 + 1
FIRST_ROW_2 = r
for rep in REPS:
    name = rep["name"]
    LOG = IMPORT_META[name]["log_sheet"]
    log_rows = IMPORT_META[name]["log_rows"]
    date_rng = f"{LOG}!$A$1:$A${log_rows}"
    type_rng = f"{LOG}!$C$1:$C${log_rows}"
    # Active Leads (Week/Month) - same live-status definition as "Today" above: counts by
    # CURRENT status (excludes Dead Lead and Potential Lead), not the historical Outreach
    # log event, pulled from the Book import tab instead of the Activity Log import tab.
    BK_wm = IMPORT_META[name]["book_sheet"]
    book_rows_wm = IMPORT_META[name]["book_rows"]
    status_rng_wm = f"{BK_wm}!$E$1:$E${book_rows_wm}"
    dateassigned_rng_wm = f"{BK_wm}!$H$1:$H${book_rows_wm}"
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=1).font = BOLD
    col = 2
    for start_expr in (week_start, month_start):
        f_outreach = (
            f'=COUNTIFS({dateassigned_rng_wm},">="&{start_expr},{dateassigned_rng_wm},"<="&TODAY(),'
            f'{status_rng_wm},"<>Dead Lead",{status_rng_wm},"<>Potential Lead")'
        )
        ws.cell(row=r, column=col, value=f_outreach)
        col += 1
        # VIP Transfers = VIP Team hand-off OR Status change into VIP Transferred - same
        # combined definition used on the Total/Active List blocks above and on the Master.
        f_vip = (
            f'=COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),{type_rng},"VIP Transfer")'
            f'+COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),{type_rng},"*-> VIP Transferred")'
        )
        ws.cell(row=r, column=col, value=f_vip)
        col += 1
        f_ftd = (
            f'=COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),'
            f'{type_rng},"*-> First Deposit")'
        )
        ws.cell(row=r, column=col, value=f_ftd)
        col += 1
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER
        if c > 1:
            ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=11, color=NAVY)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    r += 1
LAST_ROW_2 = r - 1
print("executive dashboard ok:", FIRST_ROW_1, LAST_ROW_1, "|", FIRST_ROW_2, LAST_ROW_2)

# ============================================================
# Reorder tabs: Read Me first, then the main working tabs in the order described on the
# Read Me page, then every "_Import ..." helper tab pushed to the back (still visible -
# nothing hidden - just out of the way since they're not meant to be read directly).
# ============================================================
_front_order = [n for n in ["Read Me", "Master Player DB", "Master FTD List", "VIP Pipeline",
                             "Overdue Follow-Ups", "Executive Dashboard", "Team & KPI Targets"]
                if n in wb.sheetnames]
_remaining = [n for n in wb.sheetnames if n not in _front_order]
wb._sheets = [wb[n] for n in _front_order + _remaining]
wb.active = 0

wb.save("Daily_Gamba_Master_Dashboard_v24.xlsx")
print("saved: Daily_Gamba_Master_Dashboard_v24.xlsx")
