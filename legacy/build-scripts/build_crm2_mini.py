from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import date

FONT_NAME = "Arial"
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
NORMAL = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, size=10, bold=True)
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF")
GREEN_FONT = Font(name=FONT_NAME, size=10, color="008000")
GREY_FONT = Font(name=FONT_NAME, size=10, color="999999")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE5CD")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
GREY_FILL = PatternFill("solid", fgColor="EFEFEF")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
BLACK_FILL = PatternFill("solid", fgColor="D9D9D9")
DIVIDER_FILL = PatternFill("solid", fgColor="C9DAF8")

wb = Workbook()

def style_header_row(ws, row, ncols, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ============================================================
# SHARED CONFIG
# ============================================================
# kpi dict = the 4 official company-wide KPI targets.
# "outreach" is a static reference number (cold reachouts don't need to be logged -
# there's nothing to compare it against, so no formula tracks it).
# "active_leads" is the real tracked daily target: how many NEW players get logged
# into the Book each day (reuses the existing "Outreach" Activity Log event under the
# hood - that already fires the moment a Player Handle is typed in for the first time -
# just relabeled here to match the "Active Leads" terminology).
# "vip_transfers" / "ftd" are daily counts, same as before.
REPS = [
    {"name": "Yuri", "code": "YU", "role": "Acquisition",
     "platforms": "Discord, Telegram, Twitter, Instagram",
     "default_source": "Discord", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Tuna", "code": "TU", "role": "Acquisition", "platforms": "Instagram",
     "default_source": "Instagram", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Plat", "code": "PL", "role": "Acquisition, Stream Help",
     "platforms": "Discord, Telegram, Twitter", "default_source": "", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Chella", "code": "CH", "role": "Acquisition", "platforms": "Discord, Telegram, Twitter",
     "default_source": "Discord", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Moneyheist", "code": "MH", "role": "Acquisition", "platforms": "Discord",
     "default_source": "Discord", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Seb", "code": "SB", "role": "Acquisition", "platforms": "SlotEssentials",
     "default_source": "SlotEssentials", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Jordan", "code": "JD", "role": "Acquisition", "platforms": "Discord, SlotEssentials",
     "default_source": "SlotEssentials", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Seanok", "code": "SK", "role": "Acquisition", "platforms": "SlotEssentials",
     "default_source": "SlotEssentials", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Prime", "code": "IC", "role": "Manager", "platforms": "Instagram, Discord, Twitter, Telegram, SlotEssentials",
     "default_source": "", "has_kpi": True,
     "kpi": {"outreach": 20, "active_leads": 5, "vip_transfers": 1, "ftd": 1}},
    {"name": "Daily", "code": "DL", "role": "Manager", "platforms": "Instagram, Discord, Twitter, Telegram, SlotEssentials",
     "default_source": "", "has_kpi": True,
     "kpi": {"outreach": 20, "active_leads": 5, "vip_transfers": 1, "ftd": 1}},
    {"name": "Gwen", "code": "GW", "role": "VIP Team", "platforms": "Discord, Telegram, SlotEssentials",
     "default_source": "", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Miko", "code": "MK", "role": "VIP Team", "platforms": "Discord, Telegram, SlotEssentials",
     "default_source": "", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
    {"name": "Concept", "code": "CN", "role": "VIP Team", "platforms": "Discord, Telegram, SlotEssentials",
     "default_source": "", "has_kpi": True,
     "kpi": {"outreach": 100, "active_leads": 20, "vip_transfers": 3, "ftd": 1}},
]
REP_NAMES = [r["name"] for r in REPS]

# Which single rep this run builds a full Book/Daily Task/Stats/Activity Log file for.
# Change this and re-run to generate another rep's file - Team & KPI Targets always
# shows the whole roster (every rep in REPS above) for context, regardless of this value.
ACTIVE_REP = "Yuri"

# Whether this rep is measured against the 4-metric KPI framework at all - Isac/Daily
# (managers logging their own contacts, not tracked against a quota) have this off,
# which skips the whole Team & KPI Targets tab and the target-lookup formulas on Stats.
ACTIVE_REP_HAS_KPI = next(r["has_kpi"] for r in REPS if r["name"] == ACTIVE_REP)

SOURCE_CHANNELS = ["Instagram", "Discord", "Twitter", "Telegram", "SlotEssentials", "Other"]
STATUS_STAGES = ["Initial Contact", "Interested", "VIP Transferred", "KYC Complete",
                  "First Deposit", "Active",
                  "Reactivation Queue", "Dead Lead"]
STATUS_NEXTACTION = {
    "Initial Contact": "Day 1: Check Account, Check KYC, Help Deposit",
    "Interested": "Check Account, Check KYC, Help Deposit",
    "VIP Transferred": "Day 1 / Day 2 / Day 3 check-ins - urgent, help them finish KYC and lock in their first deposit. 3 attempts with still no deposit -> auto Dead Lead.",
    "KYC Complete": "Help Deposit / Confirm Deposit Pending",
    "First Deposit": "Day 3: Confirm Playing, Resolve Issues",
    "Active": "Actively playing - check in periodically, encourage continued play. Hand to the VIP Team (separate column) when warranted, or move to Reactivation Queue if they go quiet.",
    "Reactivation Queue": "Reactivation outreach - win them back",
    "Dead Lead": "Re-target (every ~30 days): reach back out, see if anything changed. Update status if it goes anywhere.",
    "Potential Lead": "Re-target (every 7 days): reach back out, see if they're ready to pick things back up.",
}
STATUS_OFFSET = {
    "Initial Contact": 1, "Interested": 1, "KYC Complete": 1, "VIP Transferred": "1/2/3",
    "First Deposit": 3, "Active": 14,
    "Reactivation Queue": 3, "Dead Lead": 30, "Potential Lead": 7,
}

# "Potential Lead" - for a player who's gone quiet after several days of daily outreach
# (Initial Contact / Interested) with no response. Switching them here drops the cadence
# from daily to every 7 days instead of giving up (Dead Lead) or continuing to grind daily
# attempts - available to every rep. Uses the exact same generic Lists-table VLOOKUP
# mechanism as every other status, so no special-casing needed anywhere else in the Book
# formulas - it's just another row in the shared Status table above.
STATUS_STAGES = STATUS_STAGES + ["Potential Lead"]

VIP_TEAM_STATUSES = ["Yes", "No"]  # dropdown options for the Transferred to VIP Team column
KYC_STATUSES = ["Not Started", "Started", "Complete", "Failed"]
DEPOSIT_STATUSES = ["No", "Pending", "Yes"]

BOOK_ROWS = 200        # capacity per rep book - real working size, not pilot-scale anymore
DAILY_TASK_ROWS = 60   # capacity per rep daily task queue
# Matches BOOK_ROWS - every Dead Lead shows up on the Reactivation block immediately
# (not just ones due today), so this needs the same ceiling as the Book itself to
# guarantee nobody silently falls off the list as the dead-lead count grows over time.
REACT_ROWS = 200
UPCOMING_ROWS = 30      # capacity for the read-only "Coming Up" preview block
QUEUE_ROWS = 10         # company-wide queue capacity (unused placeholder)

print("config ok - reps:", REP_NAMES)

# ============================================================
# SHEET: Read Me
# ============================================================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = f"Daily Gamba — Acquisition CRM ({ACTIVE_REP})"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Same structure used company-wide - every rep gets their own copy of this file."
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color="666666")

rows = [
    ("", ""),
    ("What's in this file", "header"),
    (f"{ACTIVE_REP} - Book", "— His full player database. Every player he owns, one row each. Status drives everything else automatically."),
    (f"{ACTIVE_REP} - Daily Task", "— Auto-generated: exactly who he needs to contact today. Sorted by whoever he has gone longest without talking to - that player is always row 1. He works this top to bottom, no interpretation needed."),
    (f"{ACTIVE_REP} - Stats", "— His personal scoreboard: total list vs. active list, VIP transfers, wager, and today's Active Leads/VIP Transfers/FTD vs. target." if ACTIVE_REP_HAS_KPI else "— His personal scoreboard: total list vs. active list, VIP transfers, and wager. No KPI targets since this file isn't measured against a quota."),
] + ([("Team & KPI Targets", "— His daily KPI targets (editable).")] if ACTIVE_REP_HAS_KPI else []) + [
    ("", ""),
    ("How the automation works", "header"),
    ("Status (dropdown)", "— He sets this per player: Initial Contact -> Interested -> VIP Transferred -> KYC Complete -> First Deposit -> Active. Or Reactivation Queue / Dead Lead if they go cold. This is the main funnel and is independent from the VIP Team column below."),
    ("Transferred to VIP Team (dropdown, Book tab only)", "— Separate from Status. Yes / No. Marks whether a player has been handed off to the in-house VIP team - can happen while Status stays Active (or anything else), since the two track different things at the same time. Setting it to Yes quietly starts a Day 1/7/14 check-in cadence behind the scenes (no extra columns to manage) which stops automatically after the 3rd check-in."),
    ("Health (Book tab only)", "— A glance indicator, not what decides order. Yellow = early funnel, in progress. Green = converted and paying (First Deposit / Active). Red = overdue on follow-up, or Reactivation Queue. Black = dead lead. The Daily Task tab ignores this and sorts by most overdue instead - simpler and no colors to interpret."),
    ("Next Action", "— Auto-calculated from Status. Tells him exactly what to do, no guessing."),
    ("Next Follow-Up Date", "— Auto-calculated from Status + Last Contact Date, using the Day 0/1/3/7/14/30 cadence."),
    ("Daily Task tab", "— Pulls in anyone whose Next Follow-Up Date is today or earlier, plus anyone who still has no Roobet Username (the #1 blocker to fix) - those clear once he's contacted them today, then come right back tomorrow if still no sign-up. Sorted with the longest-neglected player first."),
    ("Follow-Up Attempts (Book tab)", "— Counts how many follow-ups he's logged on a player who still hasn't signed up for Roobet. At 3 attempts, Next Action flags \"READY FOR DEAD LEAD\" as a prompt - marking the Status is still his call. Resets to 0 automatically once they sign up."),
    ("Dead Lead Reactivation (2nd block on Daily Task)", "— Every Dead Lead shows up here right away, ranked by Next Retarget Date (soonest first). A row only turns red once that date actually arrives (~30 days after last contact) - that's the signal to act. Kept separate from the main queue so dead leads never bury a fresh, hot lead."),
    ("VIP Team check-ins", "— Once he sets Transferred to VIP Team to \"Yes,\" he gets a Day 1/7/14 check-in prompt (did they deposit, are they active with the VIP team, do they need help). Stops automatically after the 3rd check-in - nothing to remember to switch off. Daily Task surfaces the player whenever either this cadence or their normal Status cadence is due."),
    ("", ""),
    ("Two things formulas can't do (need one small script)", "header"),
    ("Date Assigned", "— Ideally freezes automatically the moment a new player is added. In this pilot it's a manual date entry; I've got an Apps Script ready that automates it once we lock the structure."),
    ("Last Contact Date", "— Ideally auto-updates when he checks a player off the Daily Task list. Same story - manual for now, automatable with the same script."),
    ("", ""),
    ("Next step", "header"),
    ("Adding this rep to the Master Dashboard", "— Once this file is a real Google Sheet and shared with Claude, send the Sheet URL and it gets added to the company-wide Master Dashboard automatically - no other changes needed."),
]
r = 5
for label, val in rows:
    if val == "header":
        ws.cell(row=r, column=2, value=label).font = SUBTITLE_FONT
        r += 1
        continue
    c1 = ws.cell(row=r, column=2, value=label)
    c1.font = BOLD
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws.cell(row=r, column=3, value=val)
    c2.font = NORMAL
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

autosize(ws, {"A": 3, "B": 26, "C": 95})
for col in ["D", "E", "F", "G"]:
    ws.column_dimensions[col].width = 12

print("readme ok")

# ============================================================
# HIDDEN SHEET: Lists
# ============================================================
lst = wb.create_sheet("Lists")
lst.sheet_state = "hidden"

def write_list(col, header, items):
    lst.cell(row=1, column=col, value=header)
    for i, v in enumerate(items, start=2):
        lst.cell(row=i, column=col, value=v)

write_list(1, "Reps", REP_NAMES)
write_list(2, "SourceChannels", SOURCE_CHANNELS)
write_list(4, "KYCStatuses", KYC_STATUSES)
write_list(5, "DepositStatuses", DEPOSIT_STATUSES)

# Status mapping table: col C=Status, D=NextAction, E=OffsetDays (or CLOSED)
lst.cell(row=1, column=3, value="Status")
lst.cell(row=1, column=4, value="NextAction")
lst.cell(row=1, column=5, value="OffsetDays")
for i, s in enumerate(STATUS_STAGES, start=2):
    lst.cell(row=i, column=3, value=s)
    lst.cell(row=i, column=4, value=STATUS_NEXTACTION[s])
    lst.cell(row=i, column=5, value=STATUS_OFFSET[s])

# Buffer = extra blank rows appended to each dropdown's range in the Lists tab, so
# adding a brand-new dropdown option is just "type it in the next blank row" - no
# rebuild needed. Trade-off: the dropdown shows a handful of blank entries at the
# bottom until they're filled in.
LIST_BUFFER = 10

def col_range(col_idx, n_items, buffer=0):
    letter = get_column_letter(col_idx)
    return f"Lists!${letter}$2:${letter}${1+n_items+buffer}"

REPS_RANGE = col_range(1, len(REP_NAMES))
SOURCE_RANGE = col_range(2, len(SOURCE_CHANNELS), LIST_BUFFER)
STATUS_RANGE = col_range(3, len(STATUS_STAGES), LIST_BUFFER)
KYC_RANGE = col_range(4, len(KYC_STATUSES), LIST_BUFFER)
DEPOSIT_RANGE = col_range(5, len(DEPOSIT_STATUSES), LIST_BUFFER)
STATUS_TABLE = f"Lists!$C$2:$E${1+len(STATUS_STAGES)+LIST_BUFFER}"

print("lists ok")

# ============================================================
# SHEET: Team & KPI Targets (skipped entirely for reps not measured against a quota)
# ============================================================
if ACTIVE_REP_HAS_KPI:
    ws = wb.create_sheet("Team & KPI Targets")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Daily KPI Targets"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Blue cells are editable inputs. This is the official 4-metric KPI framework - "
                "everything else (Follow-Ups, KYC Completed, etc.) is tracked internally on the "
                "Stats tab but isn't a formal target here. This tab is personal to this file only - "
                "other reps' targets live in their own files, not here.")
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    # The official 4 targets: Daily Outreach (static, not logged - just a reference number),
    # Active Leads (the real tracked target - new players logged into the Book each day),
    # VIP Transfers, Daily FTD.
    headers = ["Name", "Role", "Platforms", "Daily Outreach\n(static, not tracked)",
               "Active Leads\nTarget (must log)", "VIP Transfers\nTarget", "Daily FTD\nTarget"]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=32)

    # Only this file's own rep - each rep's file is personal to them, not a shared roster view.
    team_rows = [
        (rep["name"], rep["role"], rep["platforms"], rep["kpi"]["outreach"],
         rep["kpi"]["active_leads"], rep["kpi"]["vip_transfers"], rep["kpi"]["ftd"])
        for rep in REPS if rep["name"] == ACTIVE_REP
    ]

    r = HEADER_ROW + 1
    TEAM_FIRST_ROW = r
    for row in team_rows:
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) if i > 3 else Alignment(horizontal="left", vertical="center", wrap_text=True)
            if i >= 4:
                if val is None:
                    cell.value = "TBD"
                    cell.font = GREY_FONT
                else:
                    cell.font = INPUT_FONT
            elif i == 1:
                cell.font = BOLD
            else:
                cell.font = NORMAL
        ws.row_dimensions[r].height = 30
        r += 1

    LAST_TEAM_ROW = r - 1

    # Plat's VIP Transfers target was given as a range (3-5) - using the low end (3) as
    # the hard pass/fail number, noted here for anyone who wonders why it's not "3-5".
    for row_idx, row in zip(range(TEAM_FIRST_ROW, LAST_TEAM_ROW + 1), team_rows):
        if row[0] == "Plat":
            ws.cell(row=row_idx, column=6).comment = Comment(
                "Target given as a range: 3-5. Using 3 (the low end) as the pass/fail number.",
                "Daily Gamba CRM")

    autosize(ws, {"A": 13, "B": 26, "C": 24, "D": 12, "E": 12, "F": 12, "G": 10})
    ws.freeze_panes = "A5"

    print("team ok", TEAM_FIRST_ROW, LAST_TEAM_ROW)
else:
    print("team kpi skipped (no quota for this rep)")

# ============================================================
# FUNCTION: build a rep's Book tab
# ============================================================
BOOK_META = {}  # rep name -> dict with row/col info for reuse by Daily Task / Stats / Master
FTD_LIST_META = {}  # rep name -> dict with row/col info for the FTD List tab

def build_book_sheet(rep):
    name = rep["name"]
    code = rep["code"]
    default_source = rep["default_source"]
    ws = wb.create_sheet(f"{name} - Book")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Book — Player Database"
    ws["A1"].font = TITLE_FONT

    headers = ["Player ID", "Player Name / Handle", "Source", "Roobet Username", "Status", "Health",
               "Priority\n(1=Urgent)", "Date\nAssigned", "Last Contact\nDate", "Next Follow-Up\nDue",
               "Next Action", "KYC Status", "Deposit Status", "Weighted\nWager", "VIP Ready", "Notes",
               "DaysSinceContact", "DueFlag", "DueRank", "Follow-Up\nAttempts",
               "ReactivationFlag", "ReactivationRank", "VIPTeamCheckinDate", "VIPTeamCheckinCount",
               "UpcomingFlag", "UpcomingRank", "VIP Transferred\nDate", "VIP Transferred\nAttempts",
               "Transferred to\nVIP Team", "StatusNextFollowUp", "VIPTeamNextFollowUp",
               "FTD Date"]
    HEADER_ROW = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=36)

    FIRST_ROW = HEADER_ROW + 1
    LAST_ROW = FIRST_ROW + BOOK_ROWS - 1

    for idx in range(BOOK_ROWS):
        r = FIRST_ROW + idx
        ws.cell(row=r, column=1, value=f'=IF($B{r}="","","{code}-"&TEXT(ROW()-{HEADER_ROW},"0000"))')
        ws.cell(row=r, column=3, value=default_source)

        # Health (F) - reflects real player state, not just timing:
        #   Red    = overdue on follow-up 3+ days, or Reactivation Queue (gone cold)
        #   Green  = converted & paying (First Deposit onward or Active)
        #   Yellow = early funnel, in progress, on schedule (New Lead / Initial Contact / etc.)
        #   Black  = dead lead
        _green_statuses = ["First Deposit", "Active"]
        _green_or = ",".join([f'$E{r}="{s}"' for s in _green_statuses])
        f_health = (
            f'=IF($E{r}="","",'
            f'IF($E{r}="Dead Lead","Black",'
            f'IF(IF(ISNUMBER($J{r}),TODAY()-$J{r}>3,FALSE),"Red",'
            f'IF($E{r}="Reactivation Queue","Red",'
            f'IF(OR({_green_or}),"Green",'
            f'"Yellow")))))'
        )
        ws.cell(row=r, column=6, value=f_health)

        # Priority (G)
        f_prio = (
            f'=IF($F{r}="","",IF($F{r}="Red",1,IF($F{r}="Yellow",2,'
            f'IF($F{r}="Green",3,4))))'
        )
        ws.cell(row=r, column=7, value=f_prio)

        # StatusNextFollowUp (AD, col 30, hidden helper) - the normal Status-driven
        # cadence off Last Contact Date. VIP Transferred keeps its own 1/2/3-day
        # checkpoint schedule (that status still exists); everything else uses the
        # flat cadence from the Lists table.
        f_status_next = (
            f'=IF($E{r}="","",'
            f'IF($E{r}="VIP Transferred",'
            f'IF($AA{r}="","",'
            f'IF(N($AB{r})>=3,"N/A - Closed",'
            f'$AA{r}+CHOOSE(N($AB{r})+1,1,2,3))),'
            f'IF($I{r}="","",'
            f'IF(VLOOKUP($E{r},{STATUS_TABLE},3,FALSE)="CLOSED","N/A - Closed",'
            f'$I{r}+VLOOKUP($E{r},{STATUS_TABLE},3,FALSE)))))'
        )
        ws.cell(row=r, column=30, value=f_status_next)

        # VIPTeamNextFollowUp (AE, col 31, hidden helper) - runs completely
        # independently of Status. Transferred to VIP Team (AC) = "Yes" starts the Day
        # 1/7/14 in-house VIP team check-in schedule, anchored on the hidden VIPTeamCheckinDate
        # (W), tracked by the hidden VIPTeamCheckinCount (X). Automatically goes quiet once all
        # 3 checkpoints are done - nothing to remember to switch off. A player can be Active (or
        # any other Status) AND mid this schedule at the same time - both are tracked in parallel.
        f_vipteam_next = (
            f'=IF(AND($AC{r}="Yes",$W{r}<>"",N($X{r})<3),'
            f'$W{r}+CHOOSE(N($X{r})+1,1,7,14),"")'
        )
        ws.cell(row=r, column=31, value=f_vipteam_next)

        # Next Follow-Up (J) - whichever of the two above is sooner drives the
        # actual due date shown/used everywhere else (Health, DueFlag, DueRank).
        f_next = (
            f'=IF($E{r}="","",'
            f'IF(AND(ISNUMBER($AD{r}),ISNUMBER($AE{r})),MIN($AD{r},$AE{r}),'
            f'IF(ISNUMBER($AE{r}),$AE{r},'
            f'$AD{r})))'
        )
        ws.cell(row=r, column=10, value=f_next)

        # Next Action (K) - if no Roobet Username yet, that's the #1 blocker: flag it first.
        # Once 3 follow-up attempts have been made with still no sign-up, flag as dead-lead-ready
        # instead (his call to actually change the Status). VIP Team check-in reminder gets
        # appended on top whenever it's running, so both pieces of guidance show together.
        f_action = (
            f'=IF($E{r}="","",'
            f'IF($E{r}="VIP Transferred",'
            f'IF(N($AB{r})>=3,'
            f'"Should be Dead Lead (3 attempts, still no deposit) - will auto-update",'
            f'"Day "&CHOOSE(N($AB{r})+1,1,2,3)&" URGENT VIP check-in - help them finish KYC and lock in first deposit"),'
            f'IF(AND($D{r}="",$E{r}<>"Dead Lead"),'
            f'IF(N($T{r})>=3,'
            f'"READY FOR DEAD LEAD (3 attempts, still no sign-up)",'
            f'"GET ROOBET SIGN-UP - "&IFERROR(VLOOKUP($E{r},{STATUS_TABLE},2,FALSE),"")),'
            f'IFERROR(VLOOKUP($E{r},{STATUS_TABLE},2,FALSE),""))))'
            f'&IF($AC{r}="Yes",IF(N($X{r})>=3,'
            f'" | VIP Team checkpoints complete - no more automatic check-ins",'
            f'" | Day "&CHOOSE(N($X{r})+1,1,7,14)&" VIP Team check-in - confirm active/depositing"),"")'
        )
        ws.cell(row=r, column=11, value=f_action)

        # VIP Ready (O)
        f_vip = f'=IF($E{r}="","",IF(AND($E{r}="Active",$F{r}="Green"),"Yes","No"))'
        ws.cell(row=r, column=15, value=f_vip)

        # Days Since Contact (Q, hidden helper, col 17) - rounded to whole days
        ws.cell(row=r, column=17, value=f'=IF($I{r}="","",ROUND(TODAY()-$I{r},0))')
        # DueFlag (R, col 18) - due if normal cadence says so, OR no Roobet Username yet
        # (missing sign-up is the #1 conversion blocker, so it always stays on the list).
        f_due = (
            f'=IF($B{r}="",0,IF(OR('
            f'AND(ISNUMBER($J{r}),$J{r}<=TODAY(),$E{r}<>"Dead Lead"),'
            f'AND($D{r}="",$E{r}<>"",$E{r}<>"Dead Lead",'
            f'OR($I{r}="",$I{r}<TODAY()))'
            f'),1,0))'
        )
        ws.cell(row=r, column=18, value=f_due)
        # DueRank (S, col 19) - THIS is what actually decides who tops the Daily Task
        # list: most days since last contact wins (ties broken by earlier row). Simple,
        # no colors to interpret - whoever he's gone longest without talking to is #1.
        f_rank = (
            f'=IF($R{r}=1,'
            f'SUMPRODUCT(($R${FIRST_ROW}:$R${LAST_ROW}=1)*'
            f'(($Q${FIRST_ROW}:$Q${LAST_ROW}>$Q{r})+'
            f'(($Q${FIRST_ROW}:$Q${LAST_ROW}=$Q{r})*(ROW($Q${FIRST_ROW}:$Q${LAST_ROW})<ROW($Q{r})))))+1,'
            f'"")'
        )
        ws.cell(row=r, column=19, value=f_rank)

        # Follow-Up Attempts (T, col 20) - plain counter, incremented by the script each time
        # Done is checked on Daily Task while Roobet Username is still blank. Visible (not
        # hidden) so he can see it building up. Starts at 0, resets to 0 once they sign up.
        ws.cell(row=r, column=20, value=f'=IF($B{r}="","",0)')

        # ReactivationFlag (U, col 21) / ReactivationRank (V, col 22) - EVERY Dead Lead
        # is flagged immediately (not just ones whose 30-day re-target date has arrived),
        # so the full dead-lead roster is always visible on the separate "Dead Lead
        # Reactivation" block on Daily Task - kept apart from the main most-overdue-first
        # queue so it never buries a fresh, hot lead. Ranked by Next Retarget Date (J)
        # ascending, so whoever's actually due soonest (or overdue) sorts to the top;
        # the block's red highlight (on the same J<TODAY() condition as the main queue)
        # is what actually tells him who needs action today vs who's just visible for
        # awareness.
        f_reactflag = f'=IF($B{r}="",0,IF($E{r}="Dead Lead",1,0))'
        ws.cell(row=r, column=21, value=f_reactflag)
        f_reactrank = (
            f'=IF($U{r}=1,'
            f'SUMPRODUCT(($U${FIRST_ROW}:$U${LAST_ROW}=1)*'
            f'(($J${FIRST_ROW}:$J${LAST_ROW}<$J{r})+'
            f'(($J${FIRST_ROW}:$J${LAST_ROW}=$J{r})*(ROW($J${FIRST_ROW}:$J${LAST_ROW})<ROW($J{r})))))+1,'
            f'"")'
        )
        ws.cell(row=r, column=22, value=f_reactrank)

        # VIPTeamCheckinDate (W, col 23, hidden) - frozen by the script the moment
        # Transferred to VIP Team first becomes "Yes"; anchors the Day 1/7/14 checkpoint
        # schedule above. VIPTeamCheckinCount (X, col 24, hidden) - counter 0-3, incremented
        # by the script each time a VIP checkpoint is completed. Both hidden - all he sees is
        # the single Transferred to VIP Team Yes/No column plus the Next Action reminder text.
        ws.cell(row=r, column=23, value=f'=IF($B{r}="","","")')
        ws.cell(row=r, column=24, value=f'=IF($B{r}="","",0)')

        # UpcomingFlag (Y, col 25) / UpcomingRank (Z, col 26) - flags players NOT yet
        # due (DueFlag=0) whose Next Follow-Up falls within the next 7 days, ranked
        # soonest-first. Feeds the read-only "Coming Up" preview on Daily Task, kept
        # separate from the actionable due list.
        f_upflag = (
            f'=IF($B{r}="",0,IF(AND($R{r}=0,$E{r}<>"Dead Lead",'
            f'ISNUMBER($J{r}),$J{r}<=TODAY()+7),1,0))'
        )
        ws.cell(row=r, column=25, value=f_upflag)
        f_uprank = (
            f'=IF($Y{r}=1,'
            f'SUMPRODUCT(($Y${FIRST_ROW}:$Y${LAST_ROW}=1)*'
            f'(($J${FIRST_ROW}:$J${LAST_ROW}<$J{r})+'
            f'(($J${FIRST_ROW}:$J${LAST_ROW}=$J{r})*(ROW($J${FIRST_ROW}:$J${LAST_ROW})<ROW($J{r})))))+1,'
            f'"")'
        )
        ws.cell(row=r, column=26, value=f_uprank)

        # VIP Transferred Date (AA, col 27) / VIP Transferred Attempts (AB, col 28) -
        # same anchor+counter pattern as VIP Transfer Date/Check-ins above, but for
        # the pre-deposit "VIP Transferred" fast-track: Day 1 / Day 2 / Day 3 urgent
        # check-ins to lock in the first deposit. 3 attempts with still no deposit ->
        # the script auto-marks Dead Lead (same idea as the Follow-Up Attempts rule).
        ws.cell(row=r, column=27, value=f'=IF($B{r}="","","")')
        ws.cell(row=r, column=28, value=f'=IF($B{r}="","",0)')

        # Transferred to VIP Team (AC, col 29) - plain input, independent of Status.
        # Blank / "Yes" / "No". Setting "Yes" starts the Day 1/7/14 in-house VIP team
        # check-in schedule (using the hidden VIPTeamCheckinDate/Count columns above),
        # which auto-closes after the 3rd check-in. A player can carry any Status at all
        # (e.g. still "Active") while this runs in parallel.

        # FTD Date (AF, col 32) - stamped directly by the script the moment a player's
        # Status is changed to First Deposit (same pattern as VIP Transfer Date / VIPFT
        # Date above - a real value, not a live lookup formula, so it's reliable and
        # never depends on the Activity Log recalculating correctly). No manual entry.
        ws.cell(row=r, column=32, value=f'=IF($B{r}="","","")')

        for c in range(1, 17):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        for c in (20, 23, 24, 27, 28, 29, 32):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        for c in (8, 9, 10, 23, 27, 32):
            ws.cell(row=r, column=c).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=8).font = INPUT_FONT
        ws.cell(row=r, column=9).font = INPUT_FONT
        ws.cell(row=r, column=14).number_format = "$#,##0;($#,##0);-"

    dv_source = DataValidation(type="list", formula1=SOURCE_RANGE, allow_blank=True)
    dv_status = DataValidation(type="list", formula1=STATUS_RANGE, allow_blank=True)
    dv_kyc = DataValidation(type="list", formula1=KYC_RANGE, allow_blank=True)
    dv_dep = DataValidation(type="list", formula1=DEPOSIT_RANGE, allow_blank=True)
    vip_team_list = ",".join(VIP_TEAM_STATUSES)
    dv_vipteam = DataValidation(type="list", formula1=f'"{vip_team_list}"', allow_blank=True)
    for dv in (dv_source, dv_status, dv_kyc, dv_dep, dv_vipteam):
        ws.add_data_validation(dv)
    dv_source.add(f"C{FIRST_ROW}:C{LAST_ROW}")
    dv_status.add(f"E{FIRST_ROW}:E{LAST_ROW}")
    dv_kyc.add(f"L{FIRST_ROW}:L{LAST_ROW}")
    dv_dep.add(f"M{FIRST_ROW}:M{LAST_ROW}")
    dv_vipteam.add(f"AC{FIRST_ROW}:AC{LAST_ROW}")

    rngh = f"F{FIRST_ROW}:F{LAST_ROW}"
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="Red"'], fill=RED_FILL))
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="Yellow"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="Green"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(rngh, FormulaRule(formula=[f'$F{FIRST_ROW}="Black"'], fill=BLACK_FILL))

    # Duplicate Player Handle warning - highlights the handle itself if the same
    # name/handle appears more than once in the Book, so the same person never
    # accidentally gets added as two separate players.
    rngb = f"B{FIRST_ROW}:B{LAST_ROW}"
    ws.conditional_formatting.add(rngb, FormulaRule(
        formula=[f'AND($B{FIRST_ROW}<>"",COUNTIF($B${FIRST_ROW}:$B${LAST_ROW},$B{FIRST_ROW})>1)'],
        fill=RED_FILL))

    # Same check on Roobet Username - catches the case where the same real person
    # got added under two different handles but the same account.
    rngd2 = f"D{FIRST_ROW}:D{LAST_ROW}"
    ws.conditional_formatting.add(rngd2, FormulaRule(
        formula=[f'AND($D{FIRST_ROW}<>"",COUNTIF($D${FIRST_ROW}:$D${LAST_ROW},$D{FIRST_ROW})>1)'],
        fill=RED_FILL))

    for col in ("Q", "R", "S", "U", "V", "W", "X", "Y", "Z", "AD", "AE"):
        ws.column_dimensions[col].hidden = True

    # Collapse secondary columns (KYC Status, Deposit Status, Weighted Wager, Notes)
    # to keep the daily working view uncluttered. One click (+) expands them.
    for col in ("L", "M", "N"):
        ws.column_dimensions[col].outline_level = 1
        ws.column_dimensions[col].hidden = True
    ws.column_dimensions["P"].outline_level = 1
    ws.column_dimensions["P"].hidden = True
    ws.sheet_properties.outlinePr.summaryRight = True

    autosize(ws, {"A": 10, "B": 20, "C": 13, "D": 16, "E": 16, "F": 11, "G": 9, "H": 11, "I": 12,
                  "J": 13, "K": 34, "L": 12, "M": 12, "N": 11, "O": 9, "P": 26, "T": 12,
                  "AA": 12, "AB": 9, "AC": 18, "AF": 13})
    ws.freeze_panes = f"C{FIRST_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:AF{LAST_ROW}"

    # ============================================================
    # Sent to VIP Team - a live list of everyone THIS rep has personally handed off
    # (Transferred to VIP Team = Yes), so he can see who he's sent over without
    # scrolling the whole Book. Purely a read-only view - the real data still lives in
    # the main table above; this just filters it. Same FILTER-not-QUERY pattern used on
    # the Master, since it's positional and doesn't choke on blank-sampled columns.
    # ============================================================
    VIP_LIST_HEADER_ROW = LAST_ROW + 3
    ws.cell(row=VIP_LIST_HEADER_ROW - 1, column=1,
            value="Sent to VIP Team - players you've personally handed off (Transferred to VIP Team = Yes).")
    ws.cell(row=VIP_LIST_HEADER_ROW - 1, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    vip_list_headers = ["Player Handle", "Status", "Health", "Last Contact Date", "Next VIP Check-in"]
    for i, h in enumerate(vip_list_headers, start=1):
        ws.cell(row=VIP_LIST_HEADER_ROW, column=i, value=h)
    style_header_row(ws, VIP_LIST_HEADER_ROW, len(vip_list_headers), height=28)

    VIP_LIST_FIRST_ROW = VIP_LIST_HEADER_ROW + 1
    ws.cell(row=VIP_LIST_FIRST_ROW, column=1, value=(
        f'=IFERROR(FILTER({{$B${FIRST_ROW}:$B${LAST_ROW},$E${FIRST_ROW}:$E${LAST_ROW},'
        f'$F${FIRST_ROW}:$F${LAST_ROW},$I${FIRST_ROW}:$I${LAST_ROW},$AE${FIRST_ROW}:$AE${LAST_ROW}}},'
        f'$AC${FIRST_ROW}:$AC${LAST_ROW}="Yes"), "")'
    ))
    VIP_LIST_LAST_ROW = VIP_LIST_FIRST_ROW + BOOK_ROWS - 1
    for rr in range(VIP_LIST_FIRST_ROW, VIP_LIST_LAST_ROW + 1):
        ws.cell(row=rr, column=4).number_format = "yyyy-mm-dd"
        ws.cell(row=rr, column=5).number_format = "yyyy-mm-dd"
    ws.freeze_panes = f"C{FIRST_ROW}"  # unchanged - just re-affirming main-table freeze still applies

    BOOK_META[name] = {"sheet": f"{name} - Book", "header_row": HEADER_ROW,
                        "first_row": FIRST_ROW, "last_row": LAST_ROW}
    print(f"book ok: {name}", FIRST_ROW, LAST_ROW)

for rep in REPS:
    if rep["name"] == ACTIVE_REP:
        build_book_sheet(rep)

# ============================================================
# FUNCTION: build a rep's FTD List tab - separate from the Book on purpose. Players stay
# on the Book as normal (First Deposit doesn't take them off active management) - this is
# just a clean, dedicated place with Roobet Username and wager info all in one spot instead
# of scrolling the full Book. Rows are appended by the script the moment Status becomes
# First Deposit (not a live FILTER), specifically so Total Wager can stay a normal,
# independently-editable manual cell - a spilled array formula can't have a manual column
# mixed into it. Weighted Wager is a snapshot at FTD time (informational); Total Wager
# starts blank for hand entry each month, with automation to follow later.
# ============================================================
def build_ftd_list_sheet(rep):
    name = rep["name"]
    ws = wb.create_sheet(f"{name} - FTD List")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s FTD List — First Time Depositors"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Added automatically the moment a player's Status is set to First Deposit on the "
                "Book. Roobet Username, FTD Date, and Weighted Wager are all pulled in "
                "automatically and read-only - nothing to fill in by hand. Players also stay on "
                "the Book as usual, this is just a clean, dedicated view. Total Weighted Wager "
                "(all players, live) is at the bottom.")
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 28

    headers = ["Player Handle", "Roobet Username", "FTD Date", "Weighted Wager"]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=32)

    FIRST_ROW = HEADER_ROW + 1
    LAST_ROW = FIRST_ROW + BOOK_ROWS - 1  # same capacity ceiling as the Book - plenty of headroom

    for r in range(FIRST_ROW, LAST_ROW + 1):
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4).number_format = "$#,##0;($#,##0);-"
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER

    # Total Weighted Wager - live sum across every player on this list, right below the
    # capacity ceiling so it never collides with a real row even at max capacity.
    total_row = LAST_ROW + 1
    ws.cell(row=total_row, column=3, value="Total Weighted Wager:")
    ws.cell(row=total_row, column=3).font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4, value=f"=SUM(D{FIRST_ROW}:D{LAST_ROW})")
    ws.cell(row=total_row, column=4).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    ws.cell(row=total_row, column=4).number_format = "$#,##0;($#,##0);-"
    ws.cell(row=total_row, column=4).border = BORDER

    ws.freeze_panes = f"A{FIRST_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:D{LAST_ROW}"
    autosize(ws, {"A": 20, "B": 20, "C": 20, "D": 18})

    FTD_LIST_META[name] = {"sheet": f"{name} - FTD List", "header_row": HEADER_ROW,
                            "first_row": FIRST_ROW, "last_row": LAST_ROW}
    print(f"ftd list ok: {name}", FIRST_ROW, LAST_ROW)

for rep in REPS:
    if rep["name"] == ACTIVE_REP:
        build_ftd_list_sheet(rep)

# ============================================================
# FUNCTION: build a rep's Daily Task tab
# ============================================================
def build_daily_task_sheet(rep):
    name = rep["name"]
    meta = BOOK_META[name]
    BK = f"'{meta['sheet']}'"
    b_first, b_last = meta["first_row"], meta["last_row"]

    ws = wb.create_sheet(f"{name} - Daily Task")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Daily Task Queue"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Auto-pulled from his Book. Anyone overdue or due today, ranked. Work top to bottom - no thinking, just execute."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    headers = ["Task Complete", "Match Row", "Player ID", "Player Name", "Due Date", "Owner", "Source",
               "Roobet Username", "Status\n(pick to update)", "Health", "Required Action",
               "Days Since\nLast Contact", "Notes"]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=32)

    FIRST_ROW = HEADER_ROW + 1
    LAST_ROW = FIRST_ROW + DAILY_TASK_ROWS - 1

    for idx in range(DAILY_TASK_ROWS):
        r = FIRST_ROW + idx
        priority = idx + 1
        ws.cell(row=r, column=1, value=False)
        ws.cell(row=r, column=2, value=f'=IFERROR(MATCH({priority},{BK}!$S${b_first}:$S${b_last},0)+{b_first}-1,"")')
        mref = f"$B{r}"
        ws.cell(row=r, column=3, value=f'=IF({mref}="","",INDEX({BK}!$A:$A,{mref}))')
        ws.cell(row=r, column=4, value=f'=IF({mref}="","",INDEX({BK}!$B:$B,{mref}))')
        ws.cell(row=r, column=5, value=f'=IF({mref}="","",INDEX({BK}!$J:$J,{mref}))')
        ws.cell(row=r, column=6, value=f'=IF({mref}="","","{name}")')
        ws.cell(row=r, column=7, value=f'=IF({mref}="","",INDEX({BK}!$C:$C,{mref}))')
        ws.cell(row=r, column=8, value=f'=IF({mref}="","",INDEX({BK}!$D:$D,{mref}))')
        ws.cell(row=r, column=9, value=f'=IF({mref}="","",INDEX({BK}!$E:$E,{mref}))')
        ws.cell(row=r, column=10, value=f'=IF({mref}="","",INDEX({BK}!$F:$F,{mref}))')
        ws.cell(row=r, column=11, value=f'=IF({mref}="","",INDEX({BK}!$K:$K,{mref}))')
        ws.cell(row=r, column=12, value=f'=IF({mref}="","",ROUND(INDEX({BK}!$Q:$Q,{mref}),0))')
        ws.cell(row=r, column=13, value=f'=IF({mref}="","",INDEX({BK}!$P:$P,{mref}))')
        for c in range(1, 14):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=12).number_format = '0" days"'

    dv_done = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_done)
    dv_done.add(f"A{FIRST_ROW}:A{LAST_ROW}")

    # Status is editable right here - picking a new value pushes it back to the Book,
    # stamps Last Contact Date, and logs the follow-up, same as checking Done.
    dv_status_dt = DataValidation(type="list", formula1=STATUS_RANGE, allow_blank=True)
    ws.add_data_validation(dv_status_dt)
    dv_status_dt.add(f"I{FIRST_ROW}:I{LAST_ROW}")

    rng = f"J{FIRST_ROW}:J{LAST_ROW}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$J{FIRST_ROW}="Red"'], fill=RED_FILL))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$J{FIRST_ROW}="Yellow"'], fill=YELLOW_FILL))

    # Overdue highlight: red across the WHOLE row when Due Date is strictly before
    # today - meaning this one was already due yesterday (or earlier) and STILL hasn't
    # been touched (checking Task Complete or changing Status would have refreshed
    # Last Contact Date and pushed it off this list). Due today only (not yet overdue)
    # stays unhighlighted here - this is specifically the "you're now late" signal.
    overdue_rng = f"A{FIRST_ROW}:M{LAST_ROW}"
    ws.conditional_formatting.add(overdue_rng, FormulaRule(
        formula=[f'AND($E{FIRST_ROW}<>"",$E{FIRST_ROW}<TODAY())'], fill=RED_FILL))

    # ============================================================
    # Second block: Dead Lead Reactivation Queue. Every Dead Lead shows up here right
    # away (full roster, always visible) - kept separate from the main most-overdue-first
    # list above so dead leads never bury a fresh, hot lead there. Ranked by Next Retarget
    # Date, soonest first; a row only turns red once that date actually arrives (~30 days
    # after last contact) - that's the at-a-glance "act on this one today" signal. Same
    # Done checkbox mechanic applies - checking it logs the attempt and resets the clock.
    # ============================================================
    REACT_HEADER_ROW = LAST_ROW + 3
    ws.cell(row=REACT_HEADER_ROW - 1, column=1,
            value="Dead Lead Reactivation - every dead lead, ranked by Next Retarget Date. Red = due now, act on it. Check Done once you've reached back out.")
    ws.cell(row=REACT_HEADER_ROW - 1, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    headers_react = list(headers)
    headers_react[4] = "Next Retarget\nDate"
    for i, h in enumerate(headers_react, start=1):
        ws.cell(row=REACT_HEADER_ROW, column=i, value=h)
    style_header_row(ws, REACT_HEADER_ROW, len(headers_react), height=32)

    REACT_FIRST_ROW = REACT_HEADER_ROW + 1
    REACT_LAST_ROW = REACT_FIRST_ROW + REACT_ROWS - 1

    for idx in range(REACT_ROWS):
        r = REACT_FIRST_ROW + idx
        priority = idx + 1
        ws.cell(row=r, column=1, value=False)
        ws.cell(row=r, column=2, value=f'=IFERROR(MATCH({priority},{BK}!$V${b_first}:$V${b_last},0)+{b_first}-1,"")')
        mref = f"$B{r}"
        ws.cell(row=r, column=3, value=f'=IF({mref}="","",INDEX({BK}!$A:$A,{mref}))')
        ws.cell(row=r, column=4, value=f'=IF({mref}="","",INDEX({BK}!$B:$B,{mref}))')
        ws.cell(row=r, column=5, value=f'=IF({mref}="","",INDEX({BK}!$J:$J,{mref}))')
        ws.cell(row=r, column=6, value=f'=IF({mref}="","","{name}")')
        ws.cell(row=r, column=7, value=f'=IF({mref}="","",INDEX({BK}!$C:$C,{mref}))')
        ws.cell(row=r, column=8, value=f'=IF({mref}="","",INDEX({BK}!$D:$D,{mref}))')
        ws.cell(row=r, column=9, value=f'=IF({mref}="","",INDEX({BK}!$E:$E,{mref}))')
        ws.cell(row=r, column=10, value=f'=IF({mref}="","",INDEX({BK}!$F:$F,{mref}))')
        ws.cell(row=r, column=11, value=f'=IF({mref}="","",INDEX({BK}!$K:$K,{mref}))')
        ws.cell(row=r, column=12, value=f'=IF({mref}="","",ROUND(INDEX({BK}!$Q:$Q,{mref}),0))')
        ws.cell(row=r, column=13, value=f'=IF({mref}="","",INDEX({BK}!$P:$P,{mref}))')
        for c in range(1, 14):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=12).number_format = '0" days"'

    dv_done.add(f"A{REACT_FIRST_ROW}:A{REACT_LAST_ROW}")
    dv_status_dt.add(f"I{REACT_FIRST_ROW}:I{REACT_LAST_ROW}")
    rng2 = f"J{REACT_FIRST_ROW}:J{REACT_LAST_ROW}"
    ws.conditional_formatting.add(rng2, FormulaRule(formula=[f'$J{REACT_FIRST_ROW}="Black"'], fill=BLACK_FILL))

    overdue_rng2 = f"A{REACT_FIRST_ROW}:M{REACT_LAST_ROW}"
    ws.conditional_formatting.add(overdue_rng2, FormulaRule(
        formula=[f'AND($E{REACT_FIRST_ROW}<>"",$E{REACT_FIRST_ROW}<TODAY())'], fill=RED_FILL))

    # ============================================================
    # Third block: Coming Up - read-only preview of players NOT due yet but
    # scheduled within the next 7 days (soonest first). Nothing to action here -
    # no Done, no editable Status - it just shows what's headed his way so the
    # main list up top never looks like a total surprise day to day.
    # ============================================================
    UP_HEADER_ROW = REACT_LAST_ROW + 3
    ws.cell(row=UP_HEADER_ROW - 1, column=1,
            value="Coming Up (next 7 days) - not due yet, just a preview. Nothing to action here.")
    ws.cell(row=UP_HEADER_ROW - 1, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    headers_up = list(headers)
    headers_up[0] = ""
    headers_up[11] = "Days Until\nDue"
    for i, h in enumerate(headers_up, start=1):
        ws.cell(row=UP_HEADER_ROW, column=i, value=h)
    style_header_row(ws, UP_HEADER_ROW, len(headers_up), height=32)

    UP_FIRST_ROW = UP_HEADER_ROW + 1
    UP_LAST_ROW = UP_FIRST_ROW + UPCOMING_ROWS - 1

    for idx in range(UPCOMING_ROWS):
        r = UP_FIRST_ROW + idx
        priority = idx + 1
        ws.cell(row=r, column=2, value=f'=IFERROR(MATCH({priority},{BK}!$Z${b_first}:$Z${b_last},0)+{b_first}-1,"")')
        mref = f"$B{r}"
        ws.cell(row=r, column=3, value=f'=IF({mref}="","",INDEX({BK}!$A:$A,{mref}))')
        ws.cell(row=r, column=4, value=f'=IF({mref}="","",INDEX({BK}!$B:$B,{mref}))')
        ws.cell(row=r, column=5, value=f'=IF({mref}="","",INDEX({BK}!$J:$J,{mref}))')
        ws.cell(row=r, column=6, value=f'=IF({mref}="","","{name}")')
        ws.cell(row=r, column=7, value=f'=IF({mref}="","",INDEX({BK}!$C:$C,{mref}))')
        ws.cell(row=r, column=8, value=f'=IF({mref}="","",INDEX({BK}!$D:$D,{mref}))')
        ws.cell(row=r, column=9, value=f'=IF({mref}="","",INDEX({BK}!$E:$E,{mref}))')
        ws.cell(row=r, column=10, value=f'=IF({mref}="","",INDEX({BK}!$F:$F,{mref}))')
        ws.cell(row=r, column=11, value=f'=IF({mref}="","",INDEX({BK}!$K:$K,{mref}))')
        ws.cell(row=r, column=12, value=f'=IF({mref}="","",ROUND(INDEX({BK}!$J:$J,{mref})-TODAY(),0))')
        ws.cell(row=r, column=13, value=f'=IF({mref}="","",INDEX({BK}!$P:$P,{mref}))')
        for c in range(2, 14):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = NORMAL
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=12).number_format = '0" days"'

    ws.column_dimensions["B"].hidden = True
    autosize(ws, {"A": 8, "C": 10, "D": 20, "E": 12, "F": 10, "G": 12, "H": 16, "I": 16,
                  "J": 10, "K": 34, "L": 12, "M": 26})
    ws.freeze_panes = f"C{FIRST_ROW}"

    print(f"daily task ok: {name}", FIRST_ROW, LAST_ROW, "| reactivation:", REACT_FIRST_ROW, REACT_LAST_ROW,
          "| coming up:", UP_FIRST_ROW, UP_LAST_ROW)

for rep in REPS:
    if rep["name"] == ACTIVE_REP:
        build_daily_task_sheet(rep)

# ============================================================
# FUNCTION: build a rep's Activity Log (dated event history, feeds Weekly/Monthly stats)
# ============================================================
def build_activity_log_sheet(rep):
    name = rep["name"]
    ws = wb.create_sheet(f"{name} - Activity Log")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Activity Log"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Auto-filled by the script - one row per event (new contact, follow-up, VIP transfer). "
                "Don't edit this by hand. This is what powers the Weekly/Monthly numbers on the Stats tab.")
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    headers = ["Date", "Rep", "Event Type", "Player ID", "Player Handle"]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers), height=24)

    autosize(ws, {"A": 14, "B": 12, "C": 16, "D": 12, "E": 20})
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    print(f"activity log ok: {name}")

# ============================================================
# FUNCTION: build a rep's Stats tab
# ============================================================
ACTIVE_STATUSES = ["Active"]

# Live count of players currently sitting in each status - purely informational (not tied
# to any KPI target/quota), so a Dead Lead or Potential Lead is visible without inflating
# the Active Leads / Active Players numbers used elsewhere on this tab.
def write_status_breakdown(ws, header_row, status_rng):
    ws.cell(row=header_row - 1, column=1,
            value="Current Status Breakdown (live count, right now - informational only, not a KPI)")
    ws.cell(row=header_row - 1, column=1).font = SUBTITLE_FONT
    bh_headers = ["Status", "Count"]
    for i, h in enumerate(bh_headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(bh_headers), height=24)
    for i, status in enumerate(STATUS_STAGES, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=1).font = Font(name=FONT_NAME, size=11)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=f'=COUNTIF({status_rng},"{status}")')
        ws.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    return header_row + len(STATUS_STAGES)

def build_stats_sheet(rep):
    name = rep["name"]
    meta = BOOK_META[name]
    BK = f"'{meta['sheet']}'"
    b_first, b_last = meta["first_row"], meta["last_row"]
    status_rng = f"{BK}!$E${b_first}:$E${b_last}"
    # Active Leads counts by CURRENT status (excludes Dead Lead and Potential Lead), not
    # by a historical Activity Log event - a lead that's since gone dead or gone quiet
    # shouldn't keep padding the numbers reps are held accountable to. Uses the Book's own
    # Date Assigned column directly (not the Activity Log) since we need each lead's LIVE
    # status, which only the Book tracks - the Activity Log is just a point-in-time record
    # of the day they were added.
    dateassigned_rng = f"{BK}!$H${b_first}:$H${b_last}"
    id_rng = f"{BK}!$B${b_first}:$B${b_last}"
    wager_rng = f"{BK}!$N${b_first}:$N${b_last}"
    dep_rng = f"{BK}!$M${b_first}:$M${b_last}"
    # VIP Transfers now = handed to the in-house VIP team (Transferred to VIP Team = Yes),
    # counted alongside the separate "VIP Transferred" status (referral sign-up event).
    vipteam_rng = f"{BK}!$AC${b_first}:$AC${b_last}"
    vip_transfer_formula = (
        f'SUMPRODUCT(COUNTIF({status_rng},"VIP Transferred"))'
        f'+SUMPRODUCT(COUNTIF({vipteam_rng},"Yes"))'
    )

    ws = wb.create_sheet(f"{name} - Stats")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{name}'s Statistics"
    ws["A1"].font = TITLE_FONT

    # Weekly / Monthly activity - pulled from the Activity Log (dated events), not the
    # Book, because the Book only stores each player's LATEST status/date, not history.
    AL = f"'{name} - Activity Log'"
    date_rng = f"{AL}!$A$5:$A$5000"
    type_rng = f"{AL}!$C$5:$C$5000"
    week_start = "(TODAY()-WEEKDAY(TODAY(),3))"
    month_start = "DATE(YEAR(TODAY()),MONTH(TODAY()),1)"

    if not rep.get("has_kpi", True):
        # Simplified mode (managers logging their own contacts, not measured against a
        # quota): skip Total List / Active List / Today-vs-Target entirely - just the
        # Week/Month totals, moved up near the top since nothing else precedes them.
        ws["A3"] = "This Week (Mon-today) / This Month (1st-today)"
        ws["A3"].font = SUBTITLE_FONT
        wm_headers = ["Period", "Active Leads", "VIP Transfers", "FTD"]
        for i, h in enumerate(wm_headers, start=1):
            ws.cell(row=4, column=i, value=h)
        style_header_row(ws, 4, len(wm_headers), height=24)

        for row_idx, (label, start_expr) in enumerate([("This Week", week_start), ("This Month", month_start)], start=5):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=1).font = BOLD
            f_leads = (
                f'=COUNTIFS({dateassigned_rng},">="&{start_expr},{dateassigned_rng},"<="&TODAY(),'
                f'{status_rng},"<>Dead Lead",{status_rng},"<>Potential Lead")'
            )
            ws.cell(row=row_idx, column=2, value=f_leads)
            f_vip = (
                f'=COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),{type_rng},"VIP Transfer")'
                f'+COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),{type_rng},"*-> VIP Transferred")'
            )
            ws.cell(row=row_idx, column=3, value=f_vip)
            f_ftd = (
                f'=COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),'
                f'{type_rng},"*-> First Deposit")'
            )
            ws.cell(row=row_idx, column=4, value=f_ftd)
            for c in range(1, 5):
                ws.cell(row=row_idx, column=c).border = BORDER
                ws.cell(row=row_idx, column=c).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

        ws["A8"] = "Build a trend chart: select A4:D6 above, then Insert > Chart in Google Sheets. Takes 10 seconds and updates itself as new weeks/months pass."
        ws["A8"].font = Font(name=FONT_NAME, size=8, italic=True, color="666666")
        ws["A8"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[8].height = 28

        write_status_breakdown(ws, 11, status_rng)

        autosize(ws, {"A": 24, "B": 24, "C": 24, "D": 22})
        print(f"stats ok (simplified): {name}")
        return

    ws["A3"] = "Total List (all players ever assigned)"
    ws["A3"].font = SUBTITLE_FONT
    tl_headers = ["Number of Contacts", "Number of VIP Transfers", "Number of Active Players", "Total Wagered (current)"]
    for i, h in enumerate(tl_headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(tl_headers), height=28)
    ws.cell(row=5, column=1, value=f'=COUNTIF({id_rng},"<>")')
    ws.cell(row=5, column=2, value=f'={vip_transfer_formula}')
    active_terms = ",".join([f'"{s}"' for s in ACTIVE_STATUSES])
    ws.cell(row=5, column=3, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{active_terms}}}))')
    ws.cell(row=5, column=4, value=f'=SUM({wager_rng})')
    ws.cell(row=5, column=4).number_format = "$#,##0;($#,##0);-"
    for c in range(1, 5):
        ws.cell(row=5, column=c).border = BORDER
        ws.cell(row=5, column=c).font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
        ws.cell(row=5, column=c).alignment = Alignment(horizontal="center")

    ws["A7"] = "Active List (currently active players only)"
    ws["A7"].font = SUBTITLE_FONT
    al_headers = ["Number of Contacts", "Number of VIP Transfers", "Number of Active Players", "Number of FTDs"]
    for i, h in enumerate(al_headers, start=1):
        ws.cell(row=8, column=i, value=h)
    style_header_row(ws, 8, len(al_headers), height=28)
    ws.cell(row=9, column=1, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{active_terms}}}))')
    ws.cell(row=9, column=2, value=f'={vip_transfer_formula}')
    ws.cell(row=9, column=3, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{active_terms}}}))')
    fdterms = ",".join([f'"{s}"' for s in (["First Deposit"] + ACTIVE_STATUSES)])
    ws.cell(row=9, column=4, value=f'=SUMPRODUCT(COUNTIF({status_rng},{{{fdterms}}}))')
    for c in range(1, 5):
        ws.cell(row=9, column=c).border = BORDER
        ws.cell(row=9, column=c).font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
        ws.cell(row=9, column=c).alignment = Alignment(horizontal="center")

    ws["A11"] = "Note: wager figures are a running total from the Weighted Wager column in his Book. A dedicated monthly/YTD wager log can be added in Phase 2 once wager data is fed in regularly."
    ws["A11"].font = Font(name=FONT_NAME, size=8, italic=True, color="666666")
    ws["A11"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[11].height = 28

    def target_lookup(col_letter):
        return (
            f'IFERROR(INDEX(\'Team & KPI Targets\'!${col_letter}$5:${col_letter}$5,'
            f'MATCH("{name}",\'Team & KPI Targets\'!$A$5:$A$5,0)),"TBD")'
        )

    # Today vs Target - the 3 metrics that are actually tracked day to day (Daily
    # Outreach is a static reference number, not logged, so it has no "actual" to compare).
    ws["A12"] = "Today vs. Target"
    ws["A12"].font = SUBTITLE_FONT
    tvt_headers = ["Active Leads Today", "Active Leads Target",
                   "VIP Transfers Today", "VIP Transfers Target", "FTD Today", "FTD Target"]
    for i, h in enumerate(tvt_headers, start=1):
        ws.cell(row=13, column=i, value=h)
    style_header_row(ws, 13, len(tvt_headers), height=24)
    ws.cell(row=14, column=1, value=(
        f'=COUNTIFS({dateassigned_rng},TODAY(),'
        f'{status_rng},"<>Dead Lead",{status_rng},"<>Potential Lead")'
    ))
    ws.cell(row=14, column=2, value=f'={target_lookup("E")}')
    ws.cell(row=14, column=3, value=(
        f'=COUNTIFS({date_rng},TODAY(),{type_rng},"VIP Transfer")'
        f'+COUNTIFS({date_rng},TODAY(),{type_rng},"*-> VIP Transferred")'
    ))
    ws.cell(row=14, column=4, value=f'={target_lookup("F")}')
    ws.cell(row=14, column=5, value=f'=COUNTIFS({date_rng},TODAY(),{type_rng},"*-> First Deposit")')
    ws.cell(row=14, column=6, value=f'={target_lookup("G")}')
    for c in range(1, 7):
        ws.cell(row=14, column=c).border = BORDER
        ws.cell(row=14, column=c).font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
        ws.cell(row=14, column=c).alignment = Alignment(horizontal="center")
    for actual_col, target_col in [(1, 2), (3, 4), (5, 6)]:
        al = get_column_letter(actual_col)
        tl = get_column_letter(target_col)
        rng = f"{al}14:{al}14"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER(${tl}14),${al}14>=${tl}14)'], fill=GREEN_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER(${tl}14),${al}14<${tl}14)'], fill=YELLOW_FILL))

    ws["A16"] = "This Week (Mon-today) / This Month (1st-today)"
    ws["A16"].font = SUBTITLE_FONT
    wm_headers = ["Period", "Active Leads", "VIP Transfers", "FTD"]
    for i, h in enumerate(wm_headers, start=1):
        ws.cell(row=17, column=i, value=h)
    style_header_row(ws, 17, len(wm_headers), height=24)

    for row_idx, (label, start_expr) in enumerate([("This Week", week_start), ("This Month", month_start)], start=18):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=1).font = BOLD
        f_leads = (
            f'=COUNTIFS({dateassigned_rng},">="&{start_expr},{dateassigned_rng},"<="&TODAY(),'
            f'{status_rng},"<>Dead Lead",{status_rng},"<>Potential Lead")'
        )
        ws.cell(row=row_idx, column=2, value=f_leads)
        # VIP Transfers = VIP Team hand-off OR Status change into VIP Transferred - same
        # combined definition used on the Total/Active List blocks above and on the Master.
        f_vip = (
            f'=COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),{type_rng},"VIP Transfer")'
            f'+COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),{type_rng},"*-> VIP Transferred")'
        )
        ws.cell(row=row_idx, column=3, value=f_vip)
        f_ftd = (
            f'=COUNTIFS({date_rng},">="&{start_expr},{date_rng},"<="&TODAY(),'
            f'{type_rng},"*-> First Deposit")'
        )
        ws.cell(row=row_idx, column=4, value=f_ftd)
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).border = BORDER
            ws.cell(row=row_idx, column=c).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
            ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

    ws["A21"] = "Build a trend chart: select A17:D19 above, then Insert > Chart in Google Sheets. Takes 10 seconds and updates itself as new weeks/months pass."
    ws["A21"].font = Font(name=FONT_NAME, size=8, italic=True, color="666666")
    ws["A21"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[21].height = 28

    # Per-player FTD Date and Weighted Wager live on the Book itself (sort/filter there
    # directly) - not duplicated here. Total Wagered (current) above already covers the
    # rolled-up total for this Stats tab.

    write_status_breakdown(ws, 24, status_rng)

    autosize(ws, {"A": 24, "B": 24, "C": 24, "D": 22, "E": 14, "F": 14})
    print(f"stats ok: {name}")

for rep in REPS:
    if rep["name"] == ACTIVE_REP:
        build_activity_log_sheet(rep)

for rep in REPS:
    if rep["name"] == ACTIVE_REP:
        build_stats_sheet(rep)

# Reorder tabs: Read Me, Lists, Team & KPI Targets, then per rep: Daily Task, Book, FTD List,
# Stats, Activity Log. Reps see their to-do list first -- that's the main thing they act on
# each day. FTD List sits right next to Book since it's just a focused view of the same data.
_front = [n for n in ["Read Me", "Team & KPI Targets", "Lists"] if n in wb.sheetnames]
_rep_order = []
_daily_task_name = None
for rep in REPS:
    if rep["name"] != ACTIVE_REP:
        continue
    dt = f'{rep["name"]} - Daily Task'
    bk = f'{rep["name"]} - Book'
    fl = f'{rep["name"]} - FTD List'
    st = f'{rep["name"]} - Stats'
    al = f'{rep["name"]} - Activity Log'
    _rep_order = [n for n in [dt, bk, fl, st, al] if n in wb.sheetnames]
    _daily_task_name = dt

_remaining = [n for n in wb.sheetnames if n not in _front and n not in _rep_order]
_final_order = _front + _rep_order + _remaining
wb._sheets = [wb[n] for n in _final_order]
if _daily_task_name and _daily_task_name in wb.sheetnames:
    wb.active = wb.sheetnames.index(_daily_task_name)

_SAVE_NAMES = {"Tuna": "Tuna_CRM_v1.xlsx", "Plat": "Plat_CRM_v3.xlsx", "Chella": "Chella_CRM_v2.xlsx", "Isac": "Isac_CRM_v3.xlsx", "Daily": "Daily_CRM_v2.xlsx"}
wb.save(_SAVE_NAMES.get(ACTIVE_REP, f"{ACTIVE_REP}_CRM_v1.xlsx"))
print(f"saved: {ACTIVE_REP} -> {_SAVE_NAMES.get(ACTIVE_REP, f'{ACTIVE_REP}_CRM_v1.xlsx')}")
