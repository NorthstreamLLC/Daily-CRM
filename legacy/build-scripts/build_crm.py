from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

FONT_NAME = "Arial"
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
NORMAL = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, size=10, bold=True)
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF")
GREEN_FONT = Font(name=FONT_NAME, size=10, color="008000")
GREY_FONT = Font(name=FONT_NAME, size=10, color="999999")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE5CD")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
GREY_FILL = PatternFill("solid", fgColor="EFEFEF")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")

wb = Workbook()

def style_header_row(ws, row, ncols, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ============================================================
# SHEET 0: READ ME
# ============================================================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["B2"] = "Concept Acquisition & Retention CRM"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Slot Essentials — Acquisition + Retention Operating System"
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color="666666")

rows = [
    ("", ""),
    ("How this workbook is organized", "header"),
    ("Team & KPI Targets", "— Roster, division, role, and daily KPI targets for every acquisition team member. Edit the blue target cells any time you adjust numbers."),
    ("Client Book", "— The master pipeline. Every lead/player any rep is working, with contact channels (email, Discord, Telegram, X/IG), stage, and an auto-calculated next follow-up date on the 1/3/7/30-day cadence."),
    ("Follow-Up Queue", "— Auto-generated list of who is OVERDUE or DUE TODAY across the whole team, pulled live from Client Book. Check this first thing every morning."),
    ("Daily KPI Log", "— Each rep (or the manager) logs their daily numbers here. Targets and %-to-target pull in automatically from Team & KPI Targets."),
    ("Weekly Dashboard", "— Rolls everything up: did we hit our KPIs, did it convert, did overall numbers grow. This is the sheet for your weekly team meeting."),
    ("", ""),
    ("The operating rhythm", "header"),
    ("1. Work the Follow-Up Queue every day", "— nobody's account goes quiet. Contact at Day 1, Day 3, Day 7, Day 30 until they're active and transferred to VIP."),
    ("2. Log daily numbers", "— every rep fills in the Daily KPI Log before end of day. This is non-negotiable — it is the only way to know if the KPIs are the right KPIs."),
    ("3. Retarget 3x, then mark Red", "— use email + Discord + Telegram (Client Book contact columns) before falling back to scraped/cold outreach. After 3 unanswered retargets with no response, set Stage = \"Red - Unresponsive\"."),
    ("4. Review weekly using the 4-Step framework", "— see Weekly Dashboard:"),
    ("   Step 1", "Did we actually ALL complete the KPIs?"),
    ("   Step 2", "Did it actually work? (did reachouts/follow-ups convert to transfers, KYC passes, deposits?)"),
    ("   Step 3", "Did our overall numbers grow? (active players, deposits, VIP transfers)"),
    ("   Step 4", "Adjust the targets/approach based on the answers above — log the change in the Weekly Dashboard adjustment log."),
    ("", ""),
    ("Notes on roles", "header"),
    ("Andis", "Oversees both divisions; on Acquisition, manages the sales force and personally handles the biggest accounts rather than carrying a full daily grind quota."),
    ("Acquisition reps (Tuna, Chella, Jordan, Seanok, +helper)", "Each carries a full client book and the full daily KPI quota."),
    ("Retention (Miko, Gwen)", "Own the player once transferred from Acquisition; define their retention-specific KPIs the same way once this system is running."),
]
r = 5
for label, val in rows:
    if val == "header":
        ws.cell(row=r, column=2, value=label).font = SUBTITLE_FONT
        r += 1
        continue
    c1 = ws.cell(row=r, column=2, value=label)
    c1.font = BOLD
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws.cell(row=r, column=3, value=val)
    c2.font = NORMAL
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1

autosize(ws, {"A": 3, "B": 34, "C": 90})
for col in ["D", "E", "F", "G"]:
    ws.column_dimensions[col].width = 12

print("readme ok")

# ============================================================
# HIDDEN SHEET: Lists (dropdown sources)
# ============================================================
lst = wb.create_sheet("Lists")
lst.sheet_state = "hidden"

REPS = ["Andis", "Tuna", "Chella", "Jordan", "Seanok", "New Hire (TBD)"]
SOURCE_CHANNELS = ["Instagram", "X / Twitter", "Discord", "Telegram", "Forum", "SE Contact List", "Referral", "Other"]
STAGES = ["New Lead", "Contacted - Awaiting Response", "Registered", "Deposited", "Active Player",
          "VIP Transferred", "Red - Unresponsive"]
KYC_STATUSES = ["Not Started", "Pending", "Passed", "Failed"]
YES_NO = ["Yes", "No"]

def write_list(col, header, items):
    lst.cell(row=1, column=col, value=header)
    for i, v in enumerate(items, start=2):
        lst.cell(row=i, column=col, value=v)

write_list(1, "Reps", REPS)
write_list(2, "SourceChannels", SOURCE_CHANNELS)
write_list(3, "Stages", STAGES)
write_list(4, "KYCStatuses", KYC_STATUSES)
write_list(5, "YesNo", YES_NO)

def col_range(col_idx, n_items):
    letter = get_column_letter(col_idx)
    return f"Lists!${letter}$2:${letter}${1+n_items}"

REPS_RANGE = col_range(1, len(REPS))
SOURCE_RANGE = col_range(2, len(SOURCE_CHANNELS))
STAGE_RANGE = col_range(3, len(STAGES))
KYC_RANGE = col_range(4, len(KYC_STATUSES))
YESNO_RANGE = col_range(5, len(YES_NO))

print("lists ok")

# ============================================================
# SHEET: Team & KPI Targets
# ============================================================
ws = wb.create_sheet("Team & KPI Targets")
ws.sheet_view.showGridLines = False
ws["A1"] = "Team Roster & Daily KPI Targets"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Blue cells are editable inputs. Adjust targets any week based on the Weekly Dashboard review."
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

headers = ["Division", "Name", "Role", "Primary Channel", "Daily New\nReachouts",
           "Daily Follow-Ups\n(incl. 3/7/30-day retargets)", "Daily New\nVIP Transfers",
           "Daily KYC\nPasses", "Daily\nDeposits", "Notes / Scope"]
HEADER_ROW = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header_row(ws, HEADER_ROW, len(headers), height=32)

team_rows = [
    ("Division 1 - Acquisition", "Andis", "Manager (Sales Force Oversight)", "-", None, None, None, None, None,
     "Oversees Acquisition sales force + personally manages the biggest accounts. Not on the standard daily grind quota."),
    ("Division 1 - Acquisition", "Tuna", "Instagram", "Instagram", 50, 50, 5, 5, 5, ""),
    ("Division 1 - Acquisition", "Chella", "Discord / Community", "Discord", 50, 50, 5, 5, 5, ""),
    ("Division 1 - Acquisition", "Jordan", "Discord + SE", "Discord / SE Contacts", 50, 50, 5, 5, 5, ""),
    ("Division 1 - Acquisition", "Seanok", "SE Contacts", "SE Contacts", 50, 50, 5, 5, 5, "Prioritize SE contact list (email/Discord/Telegram) over scraper."),
    ("Division 1 - Acquisition", "New Hire (TBD)", "Helper (add if volume increases)", "TBD", 50, 50, 5, 5, 5, "Placeholder row - rename once hired."),
    ("Division 2 - Retention", "Andis", "Strictly Oversees", "-", None, None, None, None, None, "Oversight only."),
    ("Division 2 - Retention", "Miko", "Manager", "-", None, None, None, None, None, "Define Retention-specific daily KPIs using this same framework once Acquisition KPIs are running."),
    ("Division 2 - Retention", "Gwen", "VIP Host", "-", None, None, None, None, None, "Owns players once transferred from Acquisition."),
]

r = HEADER_ROW + 1
TEAM_FIRST_ROW = r
for row in team_rows:
    for i, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) if i > 4 else Alignment(horizontal="left", vertical="center", wrap_text=True)
        if i in (5, 6, 7, 8, 9):
            if val is None:
                cell.value = "-"
                cell.font = GREY_FONT
            else:
                cell.font = INPUT_FONT
        elif i == 1:
            cell.font = BOLD
        else:
            cell.font = NORMAL
    ws.row_dimensions[r].height = 30
    r += 1

LAST_TEAM_ROW = r - 1
autosize(ws, {"A": 20, "B": 15, "C": 26, "D": 16, "E": 12, "F": 16, "G": 12, "H": 10, "I": 10, "J": 42})
ws.freeze_panes = "A5"

print("team ok", TEAM_FIRST_ROW, LAST_TEAM_ROW)

# ============================================================
# SHEET: Client Book
# ============================================================
ws = wb.create_sheet("Client Book")
ws.sheet_view.showGridLines = False
ws["A1"] = "Client Book — Master Acquisition Pipeline"
ws["A1"].font = TITLE_FONT

cb_headers = ["Lead ID", "Owner (Rep)", "Player Name / Handle", "Source Channel",
              "Email", "Discord Handle", "Telegram Handle", "X / IG Handle",
              "Date First\nContacted", "Contact\nAttempts", "Last Contact\nDate",
              "Stage", "KYC Status", "Deposited?", "Deposit\nAmount",
              "Next Follow-Up\nDue", "Days Until\nDue", "Follow-Up\nStatus",
              "VIP Transfer\nDate", "Notes", "DueFlag", "DueRank"]
CB_HEADER_ROW = 3
for i, h in enumerate(cb_headers, start=1):
    ws.cell(row=CB_HEADER_ROW, column=i, value=h)
style_header_row(ws, CB_HEADER_ROW, len(cb_headers), height=34)

CB_FIRST_ROW = CB_HEADER_ROW + 1
CB_LAST_ROW = CB_FIRST_ROW + 199  # 200 lead capacity

sample_leads = [
    ("Tuna", "@sarah_slots22", "Instagram", "sarah.slots@gmail.com", "", "", "@sarah_slots22",
     "2026-06-28", 1, "2026-06-29", "Contacted - Awaiting Response", "Not Started", "No", None, "First DM sent, replied once."),
    ("Seanok", "Mike DeLuca", "SE Contact List", "mdeluca88@yahoo.com", "mikedee#4471", "@mikedeluca", "",
     "2026-06-20", 3, "2026-06-27", "Registered", "Pending", "No", None, "Registered, no deposit yet - push KYC."),
    ("Chella", "vegasvibes_tt", "Discord", "", "vegasvibes#0091", "", "",
     "2026-06-01", 4, "2026-06-01", "Deposited", "Passed", "Yes", 150, "Deposited, needs handholding to Day 30 before VIP transfer."),
]

for idx in range(CB_LAST_ROW - CB_FIRST_ROW + 1):
    r = CB_FIRST_ROW + idx
    ws.cell(row=r, column=1, value=f'="CB-"&TEXT(ROW()-{CB_HEADER_ROW},"000")')
    if idx < len(sample_leads):
        (owner, name, source, email, discord, tele, xig, first_contact, attempts,
         last_contact, stage, kyc, deposited, dep_amt, notes) = sample_leads[idx]
        ws.cell(row=r, column=2, value=owner)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=source)
        ws.cell(row=r, column=5, value=email)
        ws.cell(row=r, column=6, value=discord)
        ws.cell(row=r, column=7, value=tele)
        ws.cell(row=r, column=8, value=xig)
        ws.cell(row=r, column=9, value=date.fromisoformat(first_contact))
        ws.cell(row=r, column=10, value=attempts)
        ws.cell(row=r, column=11, value=date.fromisoformat(last_contact))
        ws.cell(row=r, column=12, value=stage)
        ws.cell(row=r, column=13, value=kyc)
        ws.cell(row=r, column=14, value=deposited)
        ws.cell(row=r, column=15, value=dep_amt)
        ws.cell(row=r, column=20, value=notes)

    f_due = (
        f'=IF(OR($L{r}="VIP Transferred",$L{r}="Red - Unresponsive"),"N/A - Closed",'
        f'IF($I{r}="","",'
        f'IF($J{r}<=0,$I{r}+1,'
        f'IF($J{r}=1,$I{r}+3,'
        f'IF($J{r}=2,$I{r}+7,'
        f'IF($J{r}=3,$I{r}+30,'
        f'"REVIEW - 4+ attempts"))))))'
    )
    ws.cell(row=r, column=16, value=f_due)

    f_days = f'=IF(ISNUMBER($P{r}),$P{r}-TODAY(),"")'
    ws.cell(row=r, column=17, value=f_days)

    f_status = (
        f'=IF($I{r}="","",'
        f'IF($P{r}="N/A - Closed","Closed",'
        f'IF($P{r}="REVIEW - 4+ attempts","Review Needed",'
        f'IF($Q{r}<0,"OVERDUE",'
        f'IF($Q{r}=0,"DUE TODAY","Upcoming")))))'
    )
    ws.cell(row=r, column=18, value=f_status)

    f_flag = f'=IF(OR($R{r}="OVERDUE",$R{r}="DUE TODAY"),1,0)'
    ws.cell(row=r, column=21, value=f_flag)

    f_rank = f'=IF($U{r}=1,SUM($U${CB_FIRST_ROW}:$U{r}),"")'
    ws.cell(row=r, column=22, value=f_rank)

    for c in range(1, 21):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = NORMAL
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for c in (9, 11, 16, 19):
        ws.cell(row=r, column=c).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=15).number_format = "$#,##0;($#,##0);-"
    ws.cell(row=r, column=17).alignment = Alignment(horizontal="center")

dv_reps = DataValidation(type="list", formula1=REPS_RANGE, allow_blank=True)
dv_source = DataValidation(type="list", formula1=SOURCE_RANGE, allow_blank=True)
dv_stage = DataValidation(type="list", formula1=STAGE_RANGE, allow_blank=True)
dv_kyc = DataValidation(type="list", formula1=KYC_RANGE, allow_blank=True)
dv_yesno = DataValidation(type="list", formula1=YESNO_RANGE, allow_blank=True)
for dv in (dv_reps, dv_source, dv_stage, dv_kyc, dv_yesno):
    ws.add_data_validation(dv)
dv_reps.add(f"B{CB_FIRST_ROW}:B{CB_LAST_ROW}")
dv_source.add(f"D{CB_FIRST_ROW}:D{CB_LAST_ROW}")
dv_stage.add(f"L{CB_FIRST_ROW}:L{CB_LAST_ROW}")
dv_kyc.add(f"M{CB_FIRST_ROW}:M{CB_LAST_ROW}")
dv_yesno.add(f"N{CB_FIRST_ROW}:N{CB_LAST_ROW}")

rng = f"R{CB_FIRST_ROW}:R{CB_LAST_ROW}"
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{CB_FIRST_ROW}="OVERDUE"'], fill=RED_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{CB_FIRST_ROW}="DUE TODAY"'], fill=ORANGE_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{CB_FIRST_ROW}="Upcoming"'], fill=GREEN_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{CB_FIRST_ROW}="Review Needed"'], fill=YELLOW_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{CB_FIRST_ROW}="Closed"'], fill=GREY_FILL))

rng_stage = f"L{CB_FIRST_ROW}:L{CB_LAST_ROW}"
ws.conditional_formatting.add(rng_stage, FormulaRule(formula=[f'$L{CB_FIRST_ROW}="Red - Unresponsive"'], fill=RED_FILL, font=Font(name=FONT_NAME, color="990000", bold=True)))
ws.conditional_formatting.add(rng_stage, FormulaRule(formula=[f'$L{CB_FIRST_ROW}="VIP Transferred"'], fill=GREEN_FILL, font=Font(name=FONT_NAME, color="006100", bold=True)))

ws.column_dimensions["U"].hidden = True
ws.column_dimensions["V"].hidden = True

autosize(ws, {"A": 10, "B": 13, "C": 20, "D": 16, "E": 22, "F": 16, "G": 16, "H": 16,
              "I": 13, "J": 10, "K": 13, "L": 22, "M": 12, "N": 10, "O": 11,
              "P": 15, "Q": 10, "R": 13, "S": 13, "T": 30})
ws.freeze_panes = f"C{CB_FIRST_ROW}"
ws.auto_filter.ref = f"A{CB_HEADER_ROW}:V{CB_LAST_ROW}"

print("client book ok", CB_FIRST_ROW, CB_LAST_ROW)

# ============================================================
# SHEET: Follow-Up Queue
# ============================================================
ws = wb.create_sheet("Follow-Up Queue")
ws.sheet_view.showGridLines = False
ws["A1"] = "Follow-Up Queue — Today's Priority List"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Live pull from Client Book. Sorted by who's OVERDUE or DUE TODAY. Work top to bottom every morning."
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

fq_headers = ["Priority #", "Match Row", "Lead ID", "Owner (Rep)", "Player Name / Handle",
              "Email", "Discord Handle", "Telegram Handle", "X / IG Handle",
              "Stage", "Days Overdue\n(0 = due today)", "Follow-Up Status"]
FQ_HEADER_ROW = 4
for i, h in enumerate(fq_headers, start=1):
    ws.cell(row=FQ_HEADER_ROW, column=i, value=h)
style_header_row(ws, FQ_HEADER_ROW, len(fq_headers), height=32)

FQ_FIRST_ROW = FQ_HEADER_ROW + 1
FQ_LAST_ROW = FQ_FIRST_ROW + (CB_LAST_ROW - CB_FIRST_ROW)

CB = "'Client Book'"
for idx in range(FQ_LAST_ROW - FQ_FIRST_ROW + 1):
    r = FQ_FIRST_ROW + idx
    priority = idx + 1
    ws.cell(row=r, column=1, value=priority)
    ws.cell(row=r, column=2, value=f'=IFERROR(MATCH({priority},{CB}!$V${CB_FIRST_ROW}:$V${CB_LAST_ROW},0)+{CB_FIRST_ROW}-1,"")')
    matchrow_ref = f"$B{r}"
    ws.cell(row=r, column=3, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$A:$A,{matchrow_ref}))')
    ws.cell(row=r, column=4, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$B:$B,{matchrow_ref}))')
    ws.cell(row=r, column=5, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$C:$C,{matchrow_ref}))')
    ws.cell(row=r, column=6, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$E:$E,{matchrow_ref}))')
    ws.cell(row=r, column=7, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$F:$F,{matchrow_ref}))')
    ws.cell(row=r, column=8, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$G:$G,{matchrow_ref}))')
    ws.cell(row=r, column=9, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$H:$H,{matchrow_ref}))')
    ws.cell(row=r, column=10, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$L:$L,{matchrow_ref}))')
    ws.cell(row=r, column=11, value=f'=IF({matchrow_ref}="","",-INDEX({CB}!$Q:$Q,{matchrow_ref}))')
    ws.cell(row=r, column=12, value=f'=IF({matchrow_ref}="","",INDEX({CB}!$R:$R,{matchrow_ref}))')
    for c in range(1, 13):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = NORMAL

rng = f"L{FQ_FIRST_ROW}:L{FQ_LAST_ROW}"
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$L{FQ_FIRST_ROW}="OVERDUE"'], fill=RED_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$L{FQ_FIRST_ROW}="DUE TODAY"'], fill=ORANGE_FILL))

ws.column_dimensions["B"].hidden = True
autosize(ws, {"A": 9, "C": 10, "D": 13, "E": 20, "F": 22, "G": 16, "H": 16, "I": 16,
              "J": 22, "K": 15, "L": 13})
ws.freeze_panes = f"C{FQ_FIRST_ROW}"

print("queue ok", FQ_FIRST_ROW, FQ_LAST_ROW)

# ============================================================
# SHEET: Daily KPI Log
# ============================================================
ws = wb.create_sheet("Daily KPI Log")
ws.sheet_view.showGridLines = False
ws["A1"] = "Daily KPI Log"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Every rep logs their numbers here daily. Targets and %-to-target pull automatically from Team & KPI Targets."
ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

dk_headers = ["Date", "Rep Name", "New\nReachouts", "Follow-Ups Completed\n(incl. 3/7/30-day retargets)",
              "New VIP\nTransfers", "KYC\nPasses", "Deposits\nSecured",
              "Reachout\nTarget", "Follow-Up\nTarget", "Transfer\nTarget", "KYC\nTarget", "Deposit\nTarget",
              "Reachout\n% to Target", "Follow-Up\n% to Target", "Transfer\n% to Target",
              "KYC\n% to Target", "Deposit\n% to Target", "All KPIs\nHit?"]
DK_HEADER_ROW = 4
for i, h in enumerate(dk_headers, start=1):
    ws.cell(row=DK_HEADER_ROW, column=i, value=h)
style_header_row(ws, DK_HEADER_ROW, len(dk_headers), height=40)

DK_FIRST_ROW = DK_HEADER_ROW + 1
DK_LAST_ROW = DK_FIRST_ROW + 299

TEAM = "'Team & KPI Targets'"
sample_daily = [
    ("2026-06-29", "Tuna", 52, 50, 6, 5, 5),
    ("2026-06-29", "Chella", 41, 38, 3, 2, 3),
    ("2026-06-29", "Seanok", 50, 50, 5, 5, 5),
]

for idx in range(DK_LAST_ROW - DK_FIRST_ROW + 1):
    r = DK_FIRST_ROW + idx
    if idx < len(sample_daily):
        date_str, rep, reach, foll, trans, kyc, dep = sample_daily[idx]
        ws.cell(row=r, column=1, value=date.fromisoformat(date_str)).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=rep)
        ws.cell(row=r, column=3, value=reach)
        ws.cell(row=r, column=4, value=foll)
        ws.cell(row=r, column=5, value=trans)
        ws.cell(row=r, column=6, value=kyc)
        ws.cell(row=r, column=7, value=dep)
    else:
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"

    ws.cell(row=r, column=8, value=f'=IFERROR(VLOOKUP($B{r},{TEAM}!$B:$J,4,FALSE),"")')
    ws.cell(row=r, column=9, value=f'=IFERROR(VLOOKUP($B{r},{TEAM}!$B:$J,5,FALSE),"")')
    ws.cell(row=r, column=10, value=f'=IFERROR(VLOOKUP($B{r},{TEAM}!$B:$J,6,FALSE),"")')
    ws.cell(row=r, column=11, value=f'=IFERROR(VLOOKUP($B{r},{TEAM}!$B:$J,7,FALSE),"")')
    ws.cell(row=r, column=12, value=f'=IFERROR(VLOOKUP($B{r},{TEAM}!$B:$J,8,FALSE),"")')

    ws.cell(row=r, column=13, value=f'=IF($B{r}="","",IFERROR($C{r}/$H{r},"N/A"))')
    ws.cell(row=r, column=14, value=f'=IF($B{r}="","",IFERROR($D{r}/$I{r},"N/A"))')
    ws.cell(row=r, column=15, value=f'=IF($B{r}="","",IFERROR($E{r}/$J{r},"N/A"))')
    ws.cell(row=r, column=16, value=f'=IF($B{r}="","",IFERROR($F{r}/$K{r},"N/A"))')
    ws.cell(row=r, column=17, value=f'=IF($B{r}="","",IFERROR($G{r}/$L{r},"N/A"))')

    ws.cell(row=r, column=18, value=(
        f'=IF($B{r}="","",IF(COUNTIF($M{r}:$Q{r},"N/A")=5,"N/A - Oversight",'
        f'IF(COUNTIF($M{r}:$Q{r},"<1")=0,"Y","N")))'
    ))

    for c in range(1, 19):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = NORMAL
    for c in range(13, 18):
        ws.cell(row=r, column=c).number_format = "0%"

dv_reps2 = DataValidation(type="list", formula1=REPS_RANGE, allow_blank=True)
ws.add_data_validation(dv_reps2)
dv_reps2.add(f"B{DK_FIRST_ROW}:B{DK_LAST_ROW}")

rng = f"R{DK_FIRST_ROW}:R{DK_LAST_ROW}"
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{DK_FIRST_ROW}="Y"'], fill=GREEN_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{DK_FIRST_ROW}="N"'], fill=RED_FILL))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$R{DK_FIRST_ROW}="N/A - Oversight"'], fill=GREY_FILL))

autosize(ws, {"A": 12, "B": 13, "C": 10, "D": 12, "E": 10, "F": 9, "G": 10,
              "H": 9, "I": 9, "J": 9, "K": 8, "L": 8, "M": 10, "N": 10, "O": 10, "P": 8, "Q": 10, "R": 13})
ws.freeze_panes = f"C{DK_FIRST_ROW}"
ws.auto_filter.ref = f"A{DK_HEADER_ROW}:R{DK_LAST_ROW}"

print("daily kpi ok", DK_FIRST_ROW, DK_LAST_ROW)

# ============================================================
# SHEET: Weekly Dashboard
# ============================================================
ws = wb.create_sheet("Weekly Dashboard")
ws.sheet_view.showGridLines = False
ws["A1"] = "Weekly Dashboard — Team Meeting Review"
ws["A1"].font = TITLE_FONT

ws["A3"] = "Week Start Date (edit this):"
ws["A3"].font = BOLD
ws["B3"] = date(2026, 6, 29)
ws["B3"].number_format = "yyyy-mm-dd"
ws["B3"].font = INPUT_FONT
ws["B3"].fill = YELLOW_FILL
ws["C3"] = "Week End Date:"
ws["C3"].font = BOLD
ws["D3"] = "=$B$3+6"
ws["D3"].number_format = "yyyy-mm-dd"
ws["D3"].font = GREEN_FONT
WSTART = "$B$3"
WEND = "$D$3"

DK = "'Daily KPI Log'"
TEAM = "'Team & KPI Targets'"
DKD = f"{DK}!$A${DK_FIRST_ROW}:$A${DK_LAST_ROW}"   # date column range
DKR = f"{DK}!$B${DK_FIRST_ROW}:$B${DK_LAST_ROW}"   # rep column range
DKALLHIT = f"{DK}!$R${DK_FIRST_ROW}:$R${DK_LAST_ROW}"

def dk_col_range(col_letter):
    return f"{DK}!${col_letter}${DK_FIRST_ROW}:${col_letter}${DK_LAST_ROW}"

# --- Section: Per-rep compliance ---
ws["A5"] = "Step 1 — Did we ALL complete the KPIs this week?"
ws["A5"].font = SUBTITLE_FONT

pr_headers = ["Rep", "Reachouts\n(Actual / Wk Target)", "Follow-Ups\n(Actual / Wk Target)",
              "Transfers\n(Actual / Wk Target)", "KYC Passes\n(Actual / Wk Target)",
              "Deposits\n(Actual / Wk Target)", "Days\nLogged", "Days All\nKPIs Hit", "Compliance %"]
PR_HEADER_ROW = 6
for i, h in enumerate(pr_headers, start=1):
    ws.cell(row=PR_HEADER_ROW, column=i, value=h)
style_header_row(ws, PR_HEADER_ROW, len(pr_headers), height=34)

REPS_FOR_DASH = ["Andis", "Tuna", "Chella", "Jordan", "Seanok", "New Hire (TBD)"]
PR_FIRST_ROW = PR_HEADER_ROW + 1
actual_cols = {2: "C", 3: "D", 4: "E", 5: "F", 6: "G"}
target_vlookup_idx = {2: 4, 3: 5, 4: 6, 5: 7, 6: 8}
for i, rep in enumerate(REPS_FOR_DASH):
    r = PR_FIRST_ROW + i
    ws.cell(row=r, column=1, value=rep).font = BOLD
    for col, dkcol in actual_cols.items():
        tgt_idx = target_vlookup_idx[col]
        actual_range = dk_col_range(dkcol)
        formula = (
            f'=SUMPRODUCT(({DKR}=$A{r})*({DKD}>={WSTART})*({DKD}<={WEND})*({actual_range}))'
            f'&" / "&IFERROR(VLOOKUP($A{r},{TEAM}!$B:$J,{tgt_idx},FALSE)*7,"-")'
        )
        ws.cell(row=r, column=col, value=formula)
    ws.cell(row=r, column=7, value=f'=SUMPRODUCT(({DKR}=$A{r})*({DKD}>={WSTART})*({DKD}<={WEND}))')
    ws.cell(row=r, column=8, value=f'=SUMPRODUCT(({DKR}=$A{r})*({DKD}>={WSTART})*({DKD}<={WEND})*({DKALLHIT}="Y"))')
    ws.cell(row=r, column=9, value=f'=IFERROR($H{r}/$G{r},"N/A")')
    ws.cell(row=r, column=9).number_format = "0%"
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = BORDER
        if c not in (1, 9):
            ws.cell(row=r, column=c).font = NORMAL
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
PR_LAST_ROW = PR_FIRST_ROW + len(REPS_FOR_DASH) - 1

rngc = f"I{PR_FIRST_ROW}:I{PR_LAST_ROW}"
ws.conditional_formatting.add(rngc, CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=GREEN_FILL))
ws.conditional_formatting.add(rngc, CellIsRule(operator="lessThan", formula=["0.5"], fill=RED_FILL))

autosize(ws, {"A": 15, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 9, "H": 9, "I": 12})
print("dashboard section1 ok", PR_FIRST_ROW, PR_LAST_ROW)

# --- Section: Team totals this week ---
row = PR_LAST_ROW + 2
ws.cell(row=row, column=1, value="Team Totals This Week").font = SUBTITLE_FONT
row += 1
tt_headers = ["Total Reachouts", "Total Follow-Ups", "Total New Transfers", "Total KYC Passes", "Total Deposits"]
for i, h in enumerate(tt_headers, start=1):
    ws.cell(row=row, column=i, value=h)
style_header_row(ws, row, len(tt_headers), height=22)
TT_HEADER_ROW = row
row += 1
TT_ROW = row
dkcols = ["C", "D", "E", "F", "G"]
for i, dkcol in enumerate(dkcols, start=1):
    actual_range = dk_col_range(dkcol)
    f = f'=SUMPRODUCT(({DKD}>={WSTART})*({DKD}<={WEND})*({actual_range}))'
    ws.cell(row=TT_ROW, column=i, value=f)
    ws.cell(row=TT_ROW, column=i).border = BORDER
    ws.cell(row=TT_ROW, column=i).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    ws.cell(row=TT_ROW, column=i).alignment = Alignment(horizontal="center")

# --- Section: Pipeline snapshot ---
row = TT_ROW + 2
ws.cell(row=row, column=1, value="Pipeline Snapshot (Client Book, live)").font = SUBTITLE_FONT
row += 1
PS_HEADER_ROW = row
ps_headers = ["New Lead", "Contacted - Awaiting Response", "Registered", "Deposited",
              "Active Player", "VIP Transferred", "Red - Unresponsive"]
for i, h in enumerate(ps_headers, start=1):
    ws.cell(row=PS_HEADER_ROW, column=i, value=h)
style_header_row(ws, PS_HEADER_ROW, len(ps_headers), height=32)
row += 1
PS_ROW = row
for i, stage in enumerate(ps_headers, start=1):
    f = f'=COUNTIF({CB}!$L:$L,"{stage}")'
    ws.cell(row=PS_ROW, column=i, value=f)
    ws.cell(row=PS_ROW, column=i).border = BORDER
    ws.cell(row=PS_ROW, column=i).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    ws.cell(row=PS_ROW, column=i).alignment = Alignment(horizontal="center")

print("dashboard section2 ok", TT_HEADER_ROW, PS_ROW)

# --- Step 2: Did it actually work? (conversion) ---
row = PS_ROW + 3
ws.cell(row=row, column=1, value="Step 2 — Did it actually work? (conversion this week)").font = SUBTITLE_FONT
row += 1
S2_HEADER_ROW = row
s2_headers = ["Reachouts -> Transfers", "Reachouts -> Deposits", "Follow-Ups -> Deposits", "Notes / Interpretation"]
for i, h in enumerate(s2_headers, start=1):
    ws.cell(row=S2_HEADER_ROW, column=i, value=h)
style_header_row(ws, S2_HEADER_ROW, len(s2_headers), height=22)
row += 1
S2_ROW = row
ws.cell(row=S2_ROW, column=1, value=f'=IFERROR($C${TT_ROW}/$A${TT_ROW},"N/A")')
ws.cell(row=S2_ROW, column=2, value=f'=IFERROR($E${TT_ROW}/$A${TT_ROW},"N/A")')
ws.cell(row=S2_ROW, column=3, value=f'=IFERROR($E${TT_ROW}/$B${TT_ROW},"N/A")')
for c in (1, 2, 3):
    ws.cell(row=S2_ROW, column=c).number_format = "0.0%"
    ws.cell(row=S2_ROW, column=c).border = BORDER
    ws.cell(row=S2_ROW, column=c).alignment = Alignment(horizontal="center")
    ws.cell(row=S2_ROW, column=c).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
ws.cell(row=S2_ROW, column=4, value="Add your read on the numbers here each week.")
ws.cell(row=S2_ROW, column=4).font = INPUT_FONT
ws.cell(row=S2_ROW, column=4).fill = YELLOW_FILL
ws.cell(row=S2_ROW, column=4).border = BORDER
ws.cell(row=S2_ROW, column=4).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[S2_ROW].height = 30

# --- Step 3: Did overall numbers grow? ---
row = S2_ROW + 3
ws.cell(row=row, column=1, value="Step 3 — Did our overall numbers grow?").font = SUBTITLE_FONT
row += 1
S3_HEADER_ROW = row
s3_headers = ["Active Players — Start of Week", "Active Players — End of Week", "Net Change", "% Growth"]
for i, h in enumerate(s3_headers, start=1):
    ws.cell(row=S3_HEADER_ROW, column=i, value=h)
style_header_row(ws, S3_HEADER_ROW, len(s3_headers), height=30)
row += 1
S3_ROW = row
ws.cell(row=S3_ROW, column=1, value=0)
ws.cell(row=S3_ROW, column=2, value=0)
for c in (1, 2):
    ws.cell(row=S3_ROW, column=c).font = INPUT_FONT
    ws.cell(row=S3_ROW, column=c).fill = YELLOW_FILL
    ws.cell(row=S3_ROW, column=c).border = BORDER
    ws.cell(row=S3_ROW, column=c).alignment = Alignment(horizontal="center")
ws.cell(row=S3_ROW, column=3, value=f'=$B${S3_ROW}-$A${S3_ROW}')
ws.cell(row=S3_ROW, column=4, value=f'=IFERROR($C${S3_ROW}/$A${S3_ROW},"N/A")')
ws.cell(row=S3_ROW, column=4).number_format = "0.0%"
for c in (3, 4):
    ws.cell(row=S3_ROW, column=c).border = BORDER
    ws.cell(row=S3_ROW, column=c).font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    ws.cell(row=S3_ROW, column=c).alignment = Alignment(horizontal="center")
ws.cell(row=row + 1, column=1, value="Enter Active Players count manually each week (from your platform/BI tool) — the two yellow cells above.")
ws.cell(row=row + 1, column=1).font = Font(name=FONT_NAME, size=8, italic=True, color="666666")

# --- Step 4: Adjustment log ---
row = row + 3
ws.cell(row=row, column=1, value="Step 4 — Adjustments made based on the above").font = SUBTITLE_FONT
row += 1
S4_HEADER_ROW = row
s4_headers = ["Date", "What We Changed", "Why", "Result / Follow-Up"]
for i, h in enumerate(s4_headers, start=1):
    ws.cell(row=S4_HEADER_ROW, column=i, value=h)
style_header_row(ws, S4_HEADER_ROW, len(s4_headers), height=22)
S4_FIRST_ROW = S4_HEADER_ROW + 1
S4_LAST_ROW = S4_FIRST_ROW + 11
for r in range(S4_FIRST_ROW, S4_LAST_ROW + 1):
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = NORMAL
    ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    ws.row_dimensions[r].height = 20

autosize(ws, {"A": 24, "B": 26, "C": 24, "D": 40, "E": 14, "F": 14, "G": 14})
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 22

print("dashboard done", S4_FIRST_ROW, S4_LAST_ROW)

wb.save("Acquisition_CRM.xlsx")
print("SAVED OK")
